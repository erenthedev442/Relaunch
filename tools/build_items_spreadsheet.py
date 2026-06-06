#!/usr/bin/env python3
"""Aggregate every obtainable item across Legendary's custom content/NPCs into a
balance-tracking Excel workbook. Source = the docgen-generated progression pages
(already have resolved item names, costs, tiers, job lists)."""
import re, html, sys
from pathlib import Path

DOCS = Path(r"D:\server\.claude\worktrees\busy-johnson-c59e6e\docs\progression")
OUT  = Path(r"C:\Users\richa\OneDrive\Desktop\Legendary_Obtainable_Items.xlsx")

# ---------- markdown helpers ------------------------------------------------
def clean(cell: str) -> str:
    s = cell.replace("<br>", " — ").replace("<br/>", " — ")
    s = re.sub(r"<[^>]+>", "", s)                 # strip html tags (incl item-link <a>)
    s = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", s) # md links -> text
    s = s.replace("**", "").replace("`", "")
    s = html.unescape(s)
    return re.sub(r"\s+", " ", s).strip()

def split_row(line: str):
    return [clean(c) for c in line.strip().strip("|").split("|")]

SEP = re.compile(r"^\s*\|[\s:|-]+\|\s*$")
HDR = re.compile(r"^(#{2,4})\s+(.*)")
CUR = re.compile(r"(?:paid|cost)\s+(?:is\s+paid\s+)?in\s+([A-Za-z][A-Za-z' ]+?)\s*[.\n]", re.I)

def parse_tables(path: Path):
    """Yield (h2,h3,h4,currency_hint, header_list, [rows]) for each table."""
    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    h2 = h3 = h4 = cur = tier_ctx = ""
    i = 0
    while i < len(lines):
        ln = lines[i]
        m = HDR.match(ln)
        if m:
            lvl, txt = len(m.group(1)), clean(m.group(2))
            if lvl == 2: h2, h3, h4, cur, tier_ctx = txt, "", "", "", ""
            elif lvl == 3:
                h3, h4, cur = txt, "", ""
                if re.search(r"\btier\b", txt, re.I): tier_ctx = txt
            elif lvl == 4: h4 = txt
            i += 1; continue
        cm = CUR.search(ln)
        if cm and not ln.lstrip().startswith("|"):
            cur = cm.group(1).strip()
        if ln.lstrip().startswith("|") and i + 1 < len(lines) and SEP.match(lines[i+1]):
            header = split_row(ln)
            rows = []
            j = i + 2
            while j < len(lines) and lines[j].lstrip().startswith("|"):
                if not SEP.match(lines[j]):
                    rows.append(split_row(lines[j]))
                j += 1
            yield (h2, h3, h4, tier_ctx, cur, header, rows)
            i = j; continue
        i += 1

def num(s: str) -> str:
    return s.replace(",", "").strip()

def cur_from_h2(h2: str) -> str:
    m = re.search(r"\((Hunt Marks|[A-Z][a-z]+ Medal|Infamy|Mog Tokens)", h2)
    return m.group(1) if m else ""

# ---------- collectors ------------------------------------------------------
master = []   # Source, Tier/Group, Slot/Category, Item, Cost, Currency, Jobs/Notes
augments = [] # Catalyst, Item ID, x1, x2, x3, x4
sage = []     # Cat, Category, NM, Trophy, Catalysts
reforge_sets = []   # Job, AF, Relic, Empyrean
reforge_costs = []  # Set, base->+1, +1->+2, +2->+3
boards = []   # Board, Objective, Target, Reward
login = []    # Streak, Bonus
sources = []  # Source, Currency, How to earn / note

def g(row, n):  # safe get
    return row[n] if n < len(row) else ""

# ---- Gear Vendors ----------------------------------------------------------
p = DOCS / "gear-vendors.md"
if p.exists():
    for h2, h3, h4, tier_ctx, cur, hdr, rows in parse_tables(p):
        if hdr[:3] == ["Tier", "Currency", "How to get"]:
            for r in rows:
                sources.append([f"Gear Vendor — {g(r,0)} tier", g(r,1), g(r,2)])
        if hdr[:2] == ["Item", "Cost"]:
            tier = re.sub(r"\s*\(.*\)$", "", tier_ctx) if tier_ctx else ""
            slot = h4 or (h3 if (h3 and h3 != tier_ctx) else "")
            has_jobs = len(hdr) > 2 and hdr[2] == "Jobs"
            x0 = 3 if has_jobs else 2
            for r in rows:
                m = re.match(r"^\s*([\d,]+)\s*(.*)$", g(r,1))
                amount = num(m.group(1)) if m else g(r,1)
                currency = (m.group(2).strip() if (m and m.group(2).strip())
                            else cur or cur_from_h2(h2))
                jobs = g(r,2) if has_jobs else ""
                extra = "; ".join(f"{hdr[k]}: {g(r,k)}" for k in range(x0, len(hdr))
                                  if g(r,k) and g(r,k) not in ("—", "-"))
                master.append([h2, tier, slot, g(r,0), amount, currency, jobs, extra])

# ---- Dungeon Infamy Vendor -------------------------------------------------
p = DOCS / "dungeons.md"
if p.exists():
    for h2, h3, h4, tier_ctx, cur, hdr, rows in parse_tables(p):
        if hdr[:3] == ["Item", "Cost", "Notes"]:
            for r in rows:
                amt = re.match(r"([\d,]+)\s*(.*)", g(r,1))
                cost = num(amt.group(1)) if amt else g(r,1)
                currency = (amt.group(2).strip() if amt else "") or "Infamy"
                master.append(["Dungeon Infamy Vendor", h3, "", g(r,0), cost, currency, "", g(r,2)])

# ---- GM Home: Crafting Exchange + Mystery Mog ------------------------------
p = DOCS / "gm-home.md"
if p.exists():
    pull_costs = []
    for h2, h3, h4, tier_ctx, cur, hdr, rows in parse_tables(p):
        if hdr[:4] == ["Item", "Craft", "Hunt Marks", "Daily Cap"]:
            for r in rows:
                master.append(["Crafting Exchange", g(r,1), "", g(r,0), num(g(r,2)),
                               "Hunt Marks", "", f"Daily cap {g(r,3)}"])
        elif hdr[:2] == ["Pull type", "Cost"]:
            pull_costs = [f"{g(r,0)} = {g(r,1)}" for r in rows]
        elif hdr[:3] == ["Rarity", "Prize", "Approx. Chance"]:
            for r in rows:
                master.append(["Mystery Mog", g(r,0), "", g(r,1), "", "Mog Tokens (gacha)",
                               "", f"chance {g(r,2)}"])
    if pull_costs:
        sources.append(["Mystery Mog (gacha)", "Mog Tokens", "; ".join(pull_costs)])

# ---- Reforge ---------------------------------------------------------------
p = DOCS / "reforge.md"
if p.exists():
    for h2, h3, h4, tier_ctx, cur, hdr, rows in parse_tables(p):
        if hdr[:4] == ["Job", "AF", "Relic", "Empyrean"]:
            for r in rows:
                reforge_sets.append([g(r,0), g(r,1), g(r,2), g(r,3)])
        elif hdr and hdr[0] == "Set" and any("+1" in h for h in hdr):
            for r in rows:
                reforge_costs.append([g(r,0), g(r,1), g(r,2), g(r,3)])
        elif hdr[:4] == ["Track", "NM pool", "Currency", "Armor set"]:
            for r in rows:
                sources.append([f"Reforge — {g(r,0)}", g(r,2), f"{g(r,1)} -> {g(r,3)}"])

# ---- Augments --------------------------------------------------------------
p = DOCS / "augments.md"
if p.exists():
    for h2, h3, h4, tier_ctx, cur, hdr, rows in parse_tables(p):
        if hdr[:2] == ["Catalyst", "Item ID"]:
            for r in rows:
                augments.append([g(r,0), g(r,1), g(r,2), g(r,3), g(r,4), g(r,5)])

# ---- Augment Sage NM affinities -------------------------------------------
p = DOCS / "augment-sage.md"
if p.exists():
    for h2, h3, h4, tier_ctx, cur, hdr, rows in parse_tables(p):
        if hdr and hdr[0] in ("Cat", "Category") and "NM" in hdr:
            ni, ti = hdr.index("NM"), (hdr.index("Trophy") if "Trophy" in hdr else -1)
            ci = (hdr.index("Catalysts available") if "Catalysts available" in hdr else -1)
            for r in rows:
                sage.append([g(r,0), g(r,1) if hdr[0]=="Cat" else "", g(r,ni),
                             g(r,ti) if ti>=0 else "", g(r,ci) if ci>=0 else ""])

# ---- Boards (weekly + daily) ----------------------------------------------
for fn, label in [("weekly-hunts.md", "Weekly Hunt Board"), ("daily-board.md", "Daily Board")]:
    p = DOCS / fn
    if not p.exists(): continue
    for h2, h3, h4, tier_ctx, cur, hdr, rows in parse_tables(p):
        if hdr[:2] == ["Objective", "Target"]:
            ri = hdr.index("Reward") if "Reward" in hdr else len(hdr)-1
            si = hdr.index("Source") if "Source" in hdr else -1
            for r in rows:
                boards.append([label, g(r,0), g(r,1), g(r,ri),
                               g(r,si) if si>=0 else ""])

# ---- Login rewards ---------------------------------------------------------
p = DOCS / "login-rewards.md"
if p.exists():
    for h2, h3, h4, tier_ctx, cur, hdr, rows in parse_tables(p):
        if hdr[:1] == ["Streak"]:
            for r in rows:
                login.append([g(r,0), g(r,1)])

print(f"master={len(master)} augments={len(augments)} sage={len(sage)} "
      f"reforge_sets={len(reforge_sets)} reforge_costs={len(reforge_costs)} "
      f"boards={len(boards)} login={len(login)} sources={len(sources)}")

# ---------- workbook --------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BASE_FONT = Font(name=FONT, size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F3864")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
ALT = PatternFill("solid", fgColor="F2F5FB")

wb = Workbook()

def make_sheet(name, headers, rows, widths, title=None, numeric_cols=(), freeze_col=1):
    ws = wb.create_sheet(name)
    r0 = 1
    if title:
        ws.cell(1, 1, title).font = TITLE_FONT
        ws.cell(2, 1, f"{len(rows)} rows").font = Font(name=FONT, italic=True, size=9, color="808080")
        r0 = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r0, c, h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER
    for i, row in enumerate(rows):
        rr = r0 + 1 + i
        for c, val in enumerate(row, 1):
            cell = ws.cell(rr, c, val)
            cell.font = BASE_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=(c-1 not in numeric_cols), horizontal=("right" if (c-1) in numeric_cols else "left"))
            cell.border = BORDER
            if i % 2 == 1:
                cell.fill = ALT
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(r0 + 1, freeze_col)
    ws.auto_filter.ref = f"A{r0}:{get_column_letter(len(headers))}{r0 + len(rows)}"
    return ws

# Sort master for a readable progression view
def src_order(s):
    for k, v in [("Armor",0),("Weapons",1),("Accessory NPC",2),("Accessories",3),
                 ("Dungeon",4),("Crafting",5),("Mystery",6)]:
        if s.startswith(k): return v
    return 9
master.sort(key=lambda r: (src_order(r[0]), r[1], r[2], r[3]))

# ============================================================
# ENRICH each item with its DB stats + the role-balance SCORE.
# Replicates the EXACT weights/formula from tools/score_*.py against
# the same sql/*.sql dumps the scorers read, so scores match the
# catalogs. Armor/accessories/weapons share one stat-weight map
# (verified identical); weapons add (dmg*60/delay)*2 DPS, gated per
# role by job. Stat columns come straight from item_mods.
# ============================================================
from collections import defaultdict
SQLDIR = Path(r"D:\server\sql")
modname = {}
for ln in Path(r"D:\server\scripts\enum\mod.lua").read_text(encoding="utf-8", errors="replace").splitlines():
    mm = re.match(r"\s*([A-Z][A-Z0-9_]*)\s*=\s*(\d+),", ln)
    if mm: modname.setdefault(mm.group(1), int(mm.group(2)))

STAT_DEFS = [("STR","STR"),("DEX","DEX"),("VIT","VIT"),("AGI","AGI"),("INT","INT"),("MND","MND"),("CHR","CHR"),
    ("HP","HP"),("MP","MP"),("ACC","Acc"),("ATT","Atk"),("RACC","R.Acc"),("RATT","R.Atk"),
    ("MACC","M.Acc"),("MATT","M.Atk"),("MAGIC_DAMAGE","M.Dmg"),("DEF","Def"),("EVA","Eva"),
    ("MDEF","M.Def"),("MEVA","M.Eva"),("HASTE_GEAR","Haste"),("DUAL_WIELD","DW%"),("STORETP","StoreTP"),
    ("TP_BONUS","TPBonus"),("DOUBLE_ATTACK","DA%"),("TRIPLE_ATTACK","TA%"),("QUAD_ATTACK","QA%"),
    ("CRITHITRATE","Crit%"),("CRIT_DMG_INCREASE","CritDmg"),("FASTCAST","FastCast"),("SUBTLE_BLOW","SubtleBlow"),
    ("REFRESH","Refresh"),("REGEN","Regen"),("CURE_POTENCY","CurePot"),("SKILLCHAINBONUS","SC.Dmg"),
    ("ALL_WSDMG_ALL_HITS","WS.Dmg"),("DMGPHYS","PDT"),("DMGMAGIC","MDT"),("ENMITY","Enmity"),("SPELLINTERRUPT","SIRD")]
STAT_COLS = [(modname[n], h) for n, h in STAT_DEFS if n in modname]
STAT_HEADERS = [h for _, h in STAT_COLS]

name2id = {}
for ln in (SQLDIR/"item_basic.sql").read_text(encoding="utf-8", errors="replace").splitlines():
    mm = re.match(r"^INSERT INTO `item_basic` VALUES \((\d+),\d+,'([^']*)'", ln)
    if mm: name2id.setdefault(mm.group(2), int(mm.group(1)))

imods = {}
for fn in ("item_mods.sql", "zz_custom_naked_item_mods.sql"):
    p = SQLDIR/fn
    if not p.exists(): continue
    for ln in p.read_text(encoding="utf-8", errors="replace").splitlines():
        mm = re.match(r"^INSERT INTO `item_mods` VALUES \((\d+),(\d+),(-?\d+)\)", ln)
        if mm: imods.setdefault(int(mm.group(1)), {})[int(mm.group(2))] = int(mm.group(3))

equip = {}
for ln in (SQLDIR/"item_equipment.sql").read_text(encoding="utf-8", errors="replace").splitlines():
    mm = re.match(r"^INSERT INTO `item_equipment` VALUES \((\d+),'([^']+)',(\d+),(\d+),(\d+),", ln)
    if mm: equip[int(mm.group(1))] = {"ilvl": int(mm.group(4)), "jobs": int(mm.group(5))}

weapon_stats = {}
wp = SQLDIR/"item_weapon.sql"
if wp.exists():
    for ln in wp.read_text(encoding="utf-8", errors="replace").splitlines():
        mm = re.match(r"^INSERT INTO `item_weapon` VALUES \((\d+),'[^']+',(\d+),-?\d+,-?\d+,-?\d+,-?\d+,\d+,\d+,(\d+),(\d+),", ln)
        if mm: weapon_stats[int(mm.group(1))] = {"delay": int(mm.group(3)), "dmg": int(mm.group(4))}

item_latents = defaultdict(list)
lp = SQLDIR/"item_latents.sql"
if lp.exists():
    for ln in lp.read_text(encoding="utf-8", errors="replace").splitlines():
        mm = re.match(r"^INSERT INTO `item_latents` VALUES \((\d+),(\d+),(-?\d+),(-?\d+),(-?\d+)\)", ln)
        if mm: item_latents[int(mm.group(1))].append({"modId": int(mm.group(2)), "value": int(mm.group(3)), "latentId": int(mm.group(4))})

JOB = {n: 1 << i for i, n in enumerate(['WAR','MNK','WHM','BLM','RDM','THF','PLD','DRK','BST','BRD','RNG','SAM','NIN','DRG','SMN','BLU','COR','PUP','DNC','SCH','GEO','RUN'])}
ROLE_JOBS = {'DPS':['WAR','MNK','THF','DRK','BST','BRD','RNG','SAM','NIN','DRG','BLU','COR','DNC','PUP','RUN'],'WS':['WAR','MNK','THF','DRK','BST','BRD','RNG','SAM','NIN','DRG','BLU','COR','DNC','PUP','RUN'],'TANK':['PLD','RUN','WAR','NIN'],'CASTER':['BLM','SCH','GEO','SMN','RDM'],'HEAL':['WHM','SCH','RDM','BRD','GEO']}
ROLE_MASKS = {r: sum(JOB[j] for j in js) for r, js in ROLE_JOBS.items()}
ROLE_WEIGHTS = {'DPS':{8:2.0,9:2.0,23:1.0,25:1.5,62:5.0,73:2.0,165:4.0,259:8.0,288:6.0,302:8.0,384:0.05,289:0.5,387:-3.0,421:1.5,160:-0.03,161:-0.03},'WS':{8:2.0,9:2.0,23:1.5,25:1.5,165:2.0,421:3.0,345:0.04,840:3.0,841:2.0,570:1.0},'TANK':{10:2.0,2:0.5,3:3.0,1:1.0,63:3.0,27:3.0,29:3.0,387:-8.0,389:-8.0,160:-0.06,161:-0.06,162:-0.04,163:-0.06,164:-0.04},'CASTER':{12:2.0,13:1.0,28:3.0,30:2.0,311:4.0,170:2.0,5:0.05,6:1.0,160:-0.03,163:-0.03},'HEAL':{13:2.0,14:0.5,5:0.05,6:1.5,369:30.0,374:2.0,170:2.0,30:1.0,160:-0.03,163:-0.03}}
DD_ALWAYS_LATENTS = {7, 10, 41}
MOD_SANITY_CAP = {62:30,63:30,165:20,170:30,259:15,288:20,302:20,369:10,384:300,387:30,389:30,421:30,570:50,840:50,841:50,345:500,160:5000,161:5000,162:5000,163:5000,164:5000}
_ADDED_WEIGHTS = {  # role assignments from the scoring workbook's "Unscored on gear" tab
    'DPS':    {11:1.5,506:0.1,507:0.05,368:1.0,361:0.05,362:0.1,430:8.0,508:0.1,1039:1.0,432:0.3,954:0.1,113:0.2},
    'WS':     {11:1.5,506:0.1,507:0.05,48:1.0,175:0.01,113:0.2},
    'TANK':   {68:0.5,31:0.2,370:8.0,110:0.1,168:0.5,291:1.5,109:0.3,108:0.5,166:1.0,113:0.2},
    'CASTER': {114:0.5,115:0.5,168:0.5,113:0.2},
    'HEAL':   {168:0.5,112:0.5,519:0.5,113:0.2},
}
for _r, _w in _ADDED_WEIGHTS.items(): ROLE_WEIGHTS[_r].update(_w)
MOD_SANITY_CAP.update({31:300,507:300,175:2000,361:300,430:20})
def _clamp(mid_, val):
    cap = MOD_SANITY_CAP.get(mid_, 200)
    return cap if val > cap else (-cap if val < -cap else val)
def role_score(iid, role):
    w = ROLE_WEIGHTS[role]; s = 0.0
    for mid_, val in imods.get(iid, {}).items():
        ww = w.get(mid_)
        if ww: s += _clamp(mid_, val) * ww
    for row in item_latents.get(iid, []):
        ww = w.get(row["modId"])
        if not ww: continue
        full = role in ("DPS", "WS") and row["latentId"] in DD_ALWAYS_LATENTS
        s += _clamp(row["modId"], row["value"]) * ww * (1.0 if full else 0.5)
    ws_ = weapon_stats.get(iid)
    if ws_ and ws_["delay"] > 0:
        if role == "DPS": s += (ws_["dmg"] * 60 / ws_["delay"]) * 2.0
        elif role == "WS": s += ws_["dmg"] * 0.5
    return s
def norm(nm):
    return re.sub(r"\s+", "_", nm.lower().replace("'", "").replace("’", "").strip())

allrows = []; matched = 0
for src, tier, slot, item, cost, currency, jobs, notes in master:
    iid = name2id.get(norm(item))
    sc = ["", "", "", "", "", "", ""]   # Best Role, Score, DPS, TANK, CASTER, HEAL, WS
    ilvl = dmg = delay = ""
    statvals = [""] * len(STAT_COLS)
    if iid:
        matched += 1
        ilvl = equip.get(iid, {}).get("ilvl", "")
        ws_ = weapon_stats.get(iid)
        if ws_: dmg, delay = ws_["dmg"], ws_["delay"]
        statvals = [imods.get(iid, {}).get(mid_, "") for mid_, _ in STAT_COLS]
        if iid in equip or iid in weapon_stats:
            jb = equip.get(iid, {}).get("jobs", 0)
            per = {}
            for role in ROLE_WEIGHTS:
                if jb and (jb & ROLE_MASKS[role]) == 0: continue
                v = role_score(iid, role)
                if v > 0: per[role] = v
            if per:
                best = max(per, key=per.get)
                sc = [best, round(per[best])] + [round(per[r]) if r in per else "" for r in ("DPS","TANK","CASTER","HEAL","WS")]
    allrows.append([src, tier, slot, item, (iid or ""), ilvl, cost, currency, jobs]
                   + sc + [dmg, delay] + statvals + [notes])
print(f"enrich: matched {matched}/{len(master)} item names to DB ids "
      f"({len(STAT_COLS)} stat columns)")

ALL_H = (["Source", "Tier / Group", "Slot", "Item", "Item ID", "ilvl", "Cost", "Currency", "Jobs",
          "Best Role", "Score", "DPS", "TANK", "CASTER", "HEAL", "WS", "DMG", "Delay"] + STAT_HEADERS + ["Notes"])
ALL_W = [22, 15, 10, 28, 7, 5, 6, 14, 26, 9, 7, 6, 6, 7, 6, 6, 6, 6] + [6] * len(STAT_HEADERS) + [30]
ALL_NUM = {4, 5, 6, 10, 11, 12, 13, 14, 15, 16, 17} | set(range(18, 18 + len(STAT_HEADERS)))
make_sheet("All Items", ALL_H, allrows, ALL_W,
    title="Legendary — All Obtainable Items (stats + role-balance score)",
    numeric_cols=ALL_NUM, freeze_col=5)

make_sheet("Reforge Sets",
    ["Job", "AF set", "Relic set", "Empyrean set"], reforge_sets,
    [10, 26, 26, 26], title="Reforge — per-job armor sets (base → +3)")
if reforge_costs:
    ws = wb["Reforge Sets"]
    base = 4 + len(reforge_sets) + 3
    ws.cell(base-1, 1, "Upgrade costs (Reforge Marks)").font = TITLE_FONT
    for c, h in enumerate(["Set", "base → +1", "+1 → +2", "+2 → +3"], 1):
        cell = ws.cell(base, c, h); cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.border = BORDER
    for i, row in enumerate(reforge_costs):
        for c, v in enumerate(row, 1):
            cell = ws.cell(base+1+i, c, v); cell.font = BASE_FONT; cell.border = BORDER

make_sheet("Augment Moogle",
    ["Catalyst", "Item ID", "×1", "×2", "×3", "×4"], augments,
    [22, 9, 26, 26, 26, 26], title="Augment Moogle — catalyst → augment options",
    numeric_cols=(1,))

make_sheet("Augment Sage (NM trophies)",
    ["Cat", "Category", "NM", "Trophy", "Catalysts available"], sage,
    [6, 20, 24, 24, 40], title="Augment Sage — NM affinities → trophies/catalysts")

make_sheet("Boards (Weekly & Daily)",
    ["Board", "Objective", "Target", "Reward", "Source"], boards,
    [18, 34, 10, 22, 18], title="Weekly Hunt Board & Daily Board — objectives and rewards")

make_sheet("Login Rewards",
    ["Streak", "One-Time Bonus"], login, [12, 50], title="Login streak rewards")

# Overview sheet (first)
ov = wb.create_sheet("Overview", 0)
ov.cell(1, 1, "Legendary — Obtainable Items by Source").font = TITLE_FONT
ov.cell(2, 1, "Every custom system / NPC that hands out items, the currency it costs, and how you earn that "
              "currency. 'All Items' holds the full purchasable list; the other tabs cover augments, reforge "
              "sets, boards and login rewards.").font = Font(name=FONT, italic=True, size=9, color="808080")
counts = {}
for r in master: counts[r[0]] = counts.get(r[0], 0) + 1
for c, h in enumerate(["Source / NPC", "Currency", "How to earn the currency", "# Items", "Tab"], 1):
    cell = ov.cell(4, c, h); cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.border = BORDER
rr = 5
def ov_row(name, currency, how, n, tab):
    for c, v in enumerate([name, currency, how, n, tab], 1):
        cell = ov.cell(rr, c, v); cell.font = BASE_FONT; cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center",
                                   horizontal=("right" if c == 4 else "left"))
cur_lookup = {
    "Armor Vendor": ("Beastmen/Kindred Medals", "Trade Hunt Marks for medals at the Hunting League Seals NPC"),
    "Weapons Vendor": ("Beastmen/Kindred Medals", "Trade Hunt Marks for medals at the Hunting League Seals NPC"),
    "Accessory NPC (medal-paid jewelry)": ("Beastmen/Kindred Medals", "Trade Hunt Marks for medals"),
    "Accessories Vendor (Hunt Marks — legacy)": ("Hunt Marks", "Kill NMs via the Hunting League Spawner"),
    "Dungeon Infamy Vendor": ("Infamy", "Clear dungeons (Outer Bastion -> Eternal Throne)"),
    "Crafting Exchange": ("Hunt Marks", "Kill NMs (GM Home Crafting Exchange NPC)"),
    "Mystery Mog": ("Mog Tokens", "Gacha pulls at the Mystery Mog (GM Home)"),
}
for src in sorted(counts, key=lambda s: (src_order(s), s)):
    cl = cur_lookup.get(src, ("", ""))
    ov_row(src, cl[0], cl[1], counts[src], "All Items"); rr += 1
ov_row("Augment Moogle", "Catalyst items", "Catalysts drop from NM kills; trade to permanently augment gear", len(augments), "Augment Moogle"); rr += 1
ov_row("Reforge (AF/Relic/Empy)", "AF / Relic / Empyrean Marks", "Kill the matching NM pool; upgrade base -> +3", len(reforge_sets), "Reforge Sets"); rr += 1
ov_row("Weekly + Daily Boards", "objectives", "Complete rotating objectives for marks/bonuses", len(boards), "Boards (Weekly & Daily)"); rr += 1
ov_row("Login rewards", "—", "Log in daily / hit streak milestones", len(login), "Login Rewards"); rr += 1
for w, c in zip([36, 26, 50, 8, 22], range(1, 6)):
    ov.column_dimensions[get_column_letter(c)].width = w
ov.freeze_panes = "A5"

if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print("WROTE", OUT)
