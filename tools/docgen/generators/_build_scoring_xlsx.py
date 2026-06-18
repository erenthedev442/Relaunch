"""Build Legendary_Gear_Scoring.xlsx — scenario-analysis workbook for the gear
role-scoring weights. Weights/caps/item-values are blue inputs; role scores are
LIVE Excel formulas, so editing any weight re-scores everything (incl. the real
items on the Item Lab sheet). Role-count agnostic (reads roles from gear_finder).
"""
from __future__ import annotations
import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tools.docgen.generators import gear_finder as gf

DATA = Path(__file__).resolve().parents[3] / "docs" / "assets" / "gear-data.json"
OUT = Path(r"C:\Users\richa\OneDrive\Desktop\Legendary_Gear_Scoring.xlsx")

d = json.loads(DATA.read_text(encoding="utf-8"))
items, ML = d["items"], d["meta"]["modLabels"]
gen = d["generated_at"]
ROLES = d["meta"]["roles"]
NROLE = len(ROLES)
JOBS = d["meta"]["jobs"]
ALLJOBS = (1 << len(JOBS)) - 1
SLOTBITS = d["meta"]["slotBits"]


def lbl(mid):
    return ML.get(str(mid), gf.MOD_LABELS.get(mid, f"mod {mid}"))


def slot_str(mask):
    out = []
    for bit, name in SLOTBITS:
        if mask & (1 << bit) and name not in out:
            out.append(name)
    return "/".join(out)


def jobs_str(mask):
    if mask == ALLJOBS:
        return "All"
    return " ".join(JOBS[i] for i in range(len(JOBS)) if mask & (1 << i))


def clampv(mid, v):
    cap = gf.MOD_SANITY_CAP.get(mid, 200)
    return cap if v > cap else (-cap if v < -cap else v)


def is_real_wpn(it):
    w = it.get("w")
    if not w:
        return False
    skill, dmg, delay = w
    return delay >= 50 and skill in gf.COMBAT_SKILLS and bool(it["s"] & 0b111)


freq = Counter()
for it in items:
    for mid in {m[0] for m in it.get("m", [])} | {p[0] for p in it.get("lt", [])}:
        freq[mid] += 1

weighted = sorted({m for r in gf.ROLE_WEIGHTS.values() for m in r})
weighted.sort(key=lambda m: (-max(abs(gf.ROLE_WEIGHTS[r].get(m, 0)) for r in ROLES),
                             -freq.get(m, 0)))
unscored = sorted(((m, n) for m, n in freq.items() if m not in set(weighted)),
                  key=lambda x: -x[1])

example = next((it for it in items if it["n"] == "Skulker's Bonnet +3"), None)
ex_vals = {m[0]: m[1] for m in (example or {}).get("m", [])}
ex_name = example["n"] if example else "(example)"

# Item Lab population: obtainable items that have stats, strongest first.
lab = [it for it in items if it.get("o") and (it.get("m") or it.get("w"))]
lab.sort(key=lambda x: -(max(x["sc"]) if x.get("sc") else 0))

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BASE = Font(name=FONT, size=10)
INPUT = Font(name=FONT, size=10, color="0000FF")
LINK = Font(name=FONT, size=9, color="008000")           # green = cross-sheet link
TITLE = Font(name=FONT, bold=True, size=14, color="1F3864")
NOTE = Font(name=FONT, italic=True, size=9, color="808080")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
ALT = PatternFill("solid", fgColor="F2F5FB")
RIGHT = Alignment(horizontal="right")
SCORE_FILL = PatternFill("solid", fgColor="FFF2CC")

wb = Workbook()


def hdr(ws, row, headers, widths):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w


# ===========================================================================
# Sheet 1 — Weights
# ===========================================================================
ws = wb.active
ws.title = "Weights"
ws.cell(1, 1, "Legendary — Gear Stat Scoring Weights").font = TITLE
ws.cell(2, 1, f"Per-stat weight for each role. Blue cells are inputs you can change for scenario analysis. "
              f"Score = sum(min(stat, cap) x weight). Data snapshot {gen}.").font = NOTE
HROW = 4
CAP_COL = 3 + NROLE
GEAR_COL = 4 + NROLE
hdr(ws, HROW, ["Mod ID", "Stat"] + ROLES + ["Sanity Cap", "# Gear w/ stat"],
    [8, 22] + [8] * NROLE + [11, 14])
row_of = {}
r = HROW + 1
for m in weighted:
    row_of[m] = r
    ws.cell(r, 1, m).font = BASE
    ws.cell(r, 2, lbl(m)).font = BASE
    for ci, role in enumerate(ROLES, 3):
        w = gf.ROLE_WEIGHTS[role].get(m)
        cell = ws.cell(r, ci, w if w is not None else None)
        cell.font = INPUT if w is not None else BASE
        cell.alignment = RIGHT
    cap = ws.cell(r, CAP_COL, gf.MOD_SANITY_CAP.get(m, 200)); cap.font = INPUT; cap.alignment = RIGHT
    g = ws.cell(r, GEAR_COL, freq.get(m, 0)); g.font = BASE; g.alignment = RIGHT
    for c in range(1, GEAR_COL + 1):
        ws.cell(r, c).border = BORDER
        if (r - HROW) % 2 == 0:
            ws.cell(r, c).fill = ALT
    r += 1
ws.freeze_panes = ws.cell(HROW + 1, 1)
r += 1
ws.cell(r, 1, "Constants & rules").font = Font(name=FONT, bold=True, size=11, color="1F3864"); r += 1
const_row = {}
for label, val, note in [
    ("DPS weapon weight (DPS role)", 2.0, "DPS role: (dmg x 60 / delay) x this, for real Main/Sub/Range weapons (delay >= 50)"),
    ("WS weapon weight (WS role)", 0.5, "WS role: weapon base dmg x this (heavy/slow weapons favour WS)"),
    ("Default cap (unlisted stats)", 200, "stats not in the Sanity Cap column are clamped to +/- this"),
    ("Latent (DPS/WS engaged/TP/WS)", 1.0, "latentId in {7,10,41} count at FULL weight for DPS & WS"),
    ("Latent (all other contexts)", 0.5, "every other latent counts at HALF weight"),
]:
    ws.cell(r, 2, label).font = BASE
    c = ws.cell(r, 3, val); c.font = INPUT; c.alignment = RIGHT
    ws.cell(r, 4, note).font = NOTE
    const_row[label] = r
    r += 1
DPSW_CELL = f"$C${const_row['DPS weapon weight (DPS role)']}"
WSW_CELL = f"$C${const_row['WS weapon weight (WS role)']}"

# ===========================================================================
# Sheet 2 — Scenario (single hypothetical item)
# ===========================================================================
sc = wb.create_sheet("Scenario")
sc.cell(1, 1, "Scenario calculator").font = TITLE
sc.cell(2, 1, f"Edit an item's stat values (blue, col B) — or any weight on the Weights sheet — and the role "
              f"scores recompute. Pre-filled with {ex_name}.").font = NOTE
sc.cell(3, 1, "Note: shows raw stat-weight score for every role (real items only receive scores for roles their "
              "jobs can fill). Weapon bonuses not modeled here — see Item Lab for that.").font = NOTE
SH = 4
CAP_L = get_column_letter(CAP_COL)
hdr(sc, SH, ["Stat", "Item value"] + ROLES, [22, 11] + [10] * NROLE)
rr = SH + 1
for m in weighted:
    wrow = row_of[m]
    sc.cell(rr, 1, lbl(m)).font = BASE
    v = sc.cell(rr, 2, ex_vals.get(m, 0)); v.font = INPUT; v.alignment = RIGHT
    for ci in range(NROLE):
        col = get_column_letter(3 + ci)
        f = (f"=IF(Weights!{col}{wrow}=\"\",0,"
             f"Weights!{col}{wrow}*MAX(-Weights!${CAP_L}{wrow},MIN($B{rr},Weights!${CAP_L}{wrow})))")
        cell = sc.cell(rr, 3 + ci, f); cell.font = BASE; cell.alignment = RIGHT
    for c in range(1, 3 + NROLE):
        sc.cell(rr, c).border = BORDER
        if (rr - SH) % 2 == 0:
            sc.cell(rr, c).fill = ALT
    rr += 1
tot = rr
sc.cell(tot, 1, "ROLE SCORE").font = Font(name=FONT, bold=True, size=11)
for ci in range(NROLE):
    col = get_column_letter(3 + ci)
    cell = sc.cell(tot, 3 + ci, f"=SUM({col}{SH+1}:{col}{tot-1})")
    cell.font = Font(name=FONT, bold=True, size=11); cell.alignment = RIGHT; cell.fill = SCORE_FILL
sc.freeze_panes = sc.cell(SH + 1, 1)

# ===========================================================================
# Sheet 3 — Item Lab (live scoring of real obtainable items)
# ===========================================================================
il = wb.create_sheet("Item Lab")
il.cell(1, 1, f"Item Lab — live scoring of {len(lab)} real obtainable items").font = TITLE
il.cell(2, 1, "Change any weight on the Weights sheet and every score below recomputes. The grey weight rows "
              "(3-7) are linked from Weights. Stat columns (O+) are each item's values — edit one to test a "
              "hypothetical. Stat values are pre-clamped to the sanity caps; latents are not modeled, so a few "
              "scores may differ slightly from the in-game catalog.").font = NOTE

HEADERS = ["Item", "Slot", "Jobs", "Lv", "iLvl", "DMG", "Delay", "Wpn?"] + ROLES + ["Best", "Stats"]
STATS_COL = len(HEADERS)            # readable per-item stat list (TRUE, unclamped values)
STAT0 = len(HEADERS) + 1            # first stat-matrix column
sc0 = get_column_letter(STAT0)
scN = get_column_letter(STAT0 + len(weighted) - 1)

# Linked weight rows 3..7 (one per role), aligned under the stat columns.
for ci, role in enumerate(ROLES):
    wr = 3 + ci
    rolecol = get_column_letter(3 + ci)         # Weights C/D/E/F/G
    lab_cell = il.cell(wr, 1, f"{role} wt →")
    lab_cell.font = Font(name=FONT, italic=True, size=9, color="808080"); lab_cell.alignment = RIGHT
    for k, m in enumerate(weighted):
        c = il.cell(wr, STAT0 + k, f"=Weights!{rolecol}{row_of[m]}")
        c.font = LINK; c.number_format = "0.###"; c.alignment = RIGHT

HROW2 = 9
for c, h in enumerate(HEADERS, 1):
    cell = il.cell(HROW2, c, h); cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.border = BORDER
for k, m in enumerate(weighted):
    cell = il.cell(HROW2, STAT0 + k, lbl(m)); cell.font = HDR_FONT; cell.fill = HDR_FILL
    cell.alignment = Alignment(text_rotation=90, horizontal="center")
for c, w in enumerate([26, 9, 22, 5, 5, 6, 6, 6] + [7] * NROLE + [8, 54], 1):
    il.column_dimensions[get_column_letter(c)].width = w

ridx = HROW2 + 1
score_cols = list(range(9, 9 + NROLE))          # I.. = role score columns
for it in lab:
    R = ridx
    mv = {m[0]: m[1] for m in it.get("m", [])}
    w = it.get("w")
    il.cell(R, 1, it["n"]).font = BASE
    il.cell(R, 2, slot_str(it["s"])).font = BASE
    il.cell(R, 3, jobs_str(it["j"])).font = BASE
    il.cell(R, 4, it.get("l", "")).font = BASE
    il.cell(R, 5, it.get("il", "")).font = BASE
    il.cell(R, 6, (w[1] if w else "")).font = BASE
    il.cell(R, 7, (w[2] if w else "")).font = BASE
    il.cell(R, 8, 1 if is_real_wpn(it) else 0).font = BASE
    # stat values (pre-clamped); blank when absent so SUMPRODUCT sees 0
    for k, m in enumerate(weighted):
        if m in mv and mv[m] != 0:
            cc = il.cell(R, STAT0 + k, clampv(m, mv[m])); cc.font = INPUT; cc.number_format = "0.##"
    rng = f"${sc0}{R}:${scN}{R}"
    for ci, role in enumerate(ROLES):
        wrow = 3 + ci
        f = f"=SUMPRODUCT({rng},{sc0}${wrow}:{scN}${wrow})"
        if role == "DPS":
            f += f"+IF($H{R}=1,($F{R}*60/$G{R})*Weights!{DPSW_CELL},0)"
        elif role == "WS":
            f += f"+IF($H{R}=1,$F{R}*Weights!{WSW_CELL},0)"
        cell = il.cell(R, 9 + ci, f); cell.font = BASE; cell.number_format = "0"
        cell.alignment = RIGHT; cell.fill = SCORE_FILL
    s0, sN = get_column_letter(score_cols[0]), get_column_letter(score_cols[-1])
    best = il.cell(R, 9 + NROLE,
                   f'=CHOOSE(MATCH(MAX({s0}{R}:{sN}{R}),{s0}{R}:{sN}{R},0),'
                   + ",".join(f'"{x}"' for x in ROLES) + ")")
    best.font = Font(name=FONT, bold=True, size=10)
    parts = [f"{lbl(m[0])} {m[1]:+g}" for m in it.get("m", [])]
    if w:
        parts.append(f"DMG {w[1]}/Dly {w[2]}")
    il.cell(R, STATS_COL, ", ".join(parts)).font = BASE
    ridx += 1
il.freeze_panes = il.cell(HROW2 + 1, STAT0)   # keep name+scores + header/weights visible
il.auto_filter.ref = f"A{HROW2}:{get_column_letter(STATS_COL)}{ridx - 1}"

# ===========================================================================
# Sheet 4 — Unscored stats on gear
# ===========================================================================
us = wb.create_sheet("Unscored on gear")
us.cell(1, 1, "Stats that appear on gear but score 0 in every role").font = TITLE
us.cell(2, 1, f"{len(unscored)} of {len(freq)} distinct stats on equippable gear have no weight. "
              f"Add any to the Weights sheet to start counting it. Sorted by how many pieces carry it.").font = NOTE
UH = 4
hdr(us, UH, ["Mod ID", "Stat", "# Gear w/ stat"], [8, 30, 14])
ur = UH + 1
for m, n in unscored:
    us.cell(ur, 1, m).font = BASE
    us.cell(ur, 2, lbl(m)).font = BASE
    g = us.cell(ur, 3, n); g.font = BASE; g.alignment = RIGHT
    for c in range(1, 4):
        us.cell(ur, c).border = BORDER
        if (ur - UH) % 2 == 0:
            us.cell(ur, c).fill = ALT
    ur += 1
us.freeze_panes = us.cell(UH + 1, 1)

# ===========================================================================
# Sheet 5 — How it works
# ===========================================================================
hw = wb.create_sheet("How it works")
hw.column_dimensions["A"].width = 110
lines = [
    ("Legendary — How gear scoring works", TITLE),
    ("", NOTE),
    ("Each equippable item gets a score per role. The highest-scoring role is the item's 'best role'.", BASE),
    ("", BASE),
    ("Roles:  DPS = sustained auto-attack   |   WS = weapon-skill burst   |   TANK / CASTER / HEAL.", BASE),
    ("", BASE),
    ("Formula (per role):", Font(name=FONT, bold=True, size=11)),
    ("   score = SUM over stats of:  weight[stat] x clamp(value, -cap, +cap)", BASE),
    ("   + DPS role, real weapon:  (dmg x 60 / delay) x 2.0     [Main/Sub/Range, delay >= 50, combat skill]", BASE),
    ("   + WS role,  real weapon:  dmg x 0.5                     [so heavy/slow weapons favour WS]", BASE),
    ("   'Lower is better' stats (PDT/MDT/all damage taken) use NEGATIVE weights, so a -10% (good) scores +.", BASE),
    ("   Some mods are stored x10/x100/x10000 in the DB (skillchain, damage-taken, procs); values here are", BASE),
    ("   normalized to the real in-game number (700 -> 7%), and weights/caps are scaled to match (scores unchanged).", BASE),
    ("   latents count at half weight (full for DPS/WS latentId 7,10,41); not modeled on the Item Lab sheet.", BASE),
    ("", BASE),
    ("Sheets:", Font(name=FONT, bold=True, size=11)),
    ("   Weights   — the editable weight table (blue = inputs).", BASE),
    ("   Scenario  — one hypothetical item; type stat values, see role scores.", BASE),
    ("   Item Lab  — real obtainable items; every score is a live formula. Change a weight -> all re-score.", BASE),
    ("   Unscored on gear — stats present on gear with no weight yet.", BASE),
    ("", BASE),
    ("Where the weights live (this workbook is for analysis — real changes mean editing these):", Font(name=FONT, bold=True, size=11)),
    ("   tools/score_armor.py, tools/score_weapons.py, tools/score_accessories.py   (catalog tiers)", BASE),
    ("   tools/build_items_spreadsheet.py   (obtainable-items workbook)", BASE),
    ("   tools/docgen/generators/gear_finder.py   (Gear Finder web tool)", BASE),
    ("   All five share the same weight table; keep them in sync.", NOTE),
]
for i, (txt, fnt) in enumerate(lines, 1):
    hw.cell(i, 1, txt).font = fnt

try:
    wb.calculation.fullCalcOnLoad = True
except Exception:
    pass
OUT.parent.mkdir(parents=True, exist_ok=True)
try:
    wb.save(OUT); dest = OUT
except PermissionError:
    dest = OUT.with_name(OUT.stem + "_updated.xlsx"); wb.save(dest)
print("WROTE", dest, "| Item Lab items:", len(lab), "| weighted:", len(weighted))
