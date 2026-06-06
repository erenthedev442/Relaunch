"""Excel what-if workbook: sensitivity (instant, approximate) + scenarios (exact).

Excel can't host the server's nonlinear damage math (accuracy / attack / fSTR
caps, multi-attack EV) without reimplementing the whole engine in cell formulas
-- fragile and prone to drift. So this keeps Python as the compute engine and
Excel as the interactive front-end, in two complementary sheets:

  * "Sensitivity" -- Python computes the marginal auto-DPS / WS / rotation value
    of ONE point of each mod at your current set (by delta-perturbing the
    equipped gear through the real engine). You type hypothetical augment/gear
    amounts and Excel sums amount x per-point LIVE. First-order approximate:
    exact for linear mods, ~0 for anything you're already capped on (the useful
    "don't bother grinding this" signal), drifts only on big combined swings.

  * "Scenarios" -- you fill rows with hypothetical items/augments (an existing
    item + extra augment mods, or a brand-new item typed as a raw mod list); the
    `score` command runs each through the REAL engine and writes back exact
    auto-DPS / WS / rotation numbers and the delta vs your equipped set. This is
    the accurate verdict for a go/no-go decision.

Both pin to the same `!mystats` anchor and fixed weapon as eval / optimize.
"""
from __future__ import annotations

import re
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import engine, gear, state, targets, wsdata
from .mods import name_to_id

# ---------------------------------------------------------------------------
# Mods exposed on the Sensitivity sheet.  (label, mod name, unit hint, predicate)
# predicate(wctx) decides whether the row is relevant to this weapon setup.
# ---------------------------------------------------------------------------
def _has_off(wc) -> bool:
    return bool(wc and wc.has_offhand)


def _is_dw(wc) -> bool:
    return bool(wc and wc.has_offhand and not wc.is_h2h)


SENS_MODS: list[tuple[str, str, str, object]] = [
    ("STR",                "STR",                 "pt",            None),
    ("DEX",                "DEX",                 "pt",            None),
    ("VIT",                "VIT",                 "pt",            None),
    ("AGI",                "AGI",                 "pt",            None),
    ("Attack",             "ATT",                 "pt",            None),
    ("Accuracy",           "ACC",                 "pt",            None),
    ("Weapon DMG (main)",  "MAIN_DMG_RATING",     "pt",            None),
    ("Weapon DMG (sub)",   "SUB_DMG_RATING",      "pt",            _has_off),
    ("Double Attack",      "DOUBLE_ATTACK",       "%",             None),
    ("Triple Attack",      "TRIPLE_ATTACK",       "%",             None),
    ("Quad Attack",        "QUAD_ATTACK",         "%",             None),
    ("Crit hit rate",      "CRITHITRATE",         "%",             None),
    ("Crit damage",        "CRIT_DMG_INCREASE",   "%",             None),
    ("Haste (gear)",       "HASTE_GEAR",          "/1024 (~10=1%)", None),
    ("Dual Wield",         "DUAL_WIELD",          "%",             _is_dw),
    ("Store TP",           "STORETP",             "pt",            None),
    ("TP Bonus",           "TP_BONUS",            "pt",            None),
    ("WS dmg (all hits)",  "ALL_WSDMG_ALL_HITS",  "%",             None),
    ("WS dmg (1st hit)",   "ALL_WSDMG_FIRST_HIT", "%",             None),
    ("fTP bonus (Fotia)",  "ANY_FTP_BONUS",       "/256 (25=1pc)", None),
    ("WS accuracy",        "WSACC",               "pt",            None),
]

# Larger perturbation for stats avoids floor(STR*mult) quantization noise; the
# result is divided back to a clean per-1 marginal.
_DELTA = {"STR": 10, "DEX": 10, "VIT": 10, "AGI": 10}


# ---------------------------------------------------------------------------
# Metric helpers
# ---------------------------------------------------------------------------
@dataclass
class WhatIfCtx:
    charname: str
    target: object          # targets.Target
    ws: object | None       # wsdata.WSParams (for WS / rotation columns)
    tp: int                 # TP used for the WS / rotation columns
    main_skill: int


def metric_triplet(cs: state.CombatStats, ctx: WhatIfCtx) -> tuple[float, float, float]:
    """(auto DPS, WS damage @ctx.tp, rotation DPS) for one CombatStats."""
    auto = engine.auto_attack_dps(cs, ctx.target).dps
    if ctx.ws is None:
        return auto, 0.0, auto
    wsd = engine.ws_damage(cs, ctx.target, ctx.ws, ctx.tp).damage
    rot = engine.rotation_dps(cs, ctx.target, ctx.ws, ctx.tp).total_dps
    return auto, wsd, rot


def _build(anchor, equipped_mods, cand_mods, wctx) -> state.CombatStats:
    return state.build_combat_stats(anchor, equipped_mods, cand_mods, wctx)


def compute_sensitivity(anchor, equipped_mods, wctx, ctx: WhatIfCtx):
    """For each relevant mod: per-point marginal (auto, ws, rotation)."""
    base_cs = _build(anchor, equipped_mods, equipped_mods, wctx)
    base = metric_triplet(base_cs, ctx)
    mid = state.MID
    rows = []
    for label, key, unit, pred in SENS_MODS:
        if pred is not None and not pred(wctx):
            continue
        mod_id = mid.get(key)
        if mod_id is None:
            continue
        d = _DELTA.get(key, 1)
        cand = dict(equipped_mods)
        cand[mod_id] = cand.get(mod_id, 0) + d
        cs = _build(anchor, equipped_mods, cand, wctx)
        a, w, r = metric_triplet(cs, ctx)
        rows.append((label, unit, (a - base[0]) / d, (w - base[1]) / d,
                     (r - base[2]) / d))
    return base, rows


# ---------------------------------------------------------------------------
# Scenario parsing + scoring
# ---------------------------------------------------------------------------
_TOKEN_RE = re.compile(r"([A-Za-z_][A-Za-z0-9_ ]*?)\s*[:=+]?\s*(-?\d+)")


def parse_mod_string(s: str) -> tuple[dict[int, int], list[str]]:
    """'STR:15 ATT+20, DOUBLE_ATTACK 3' -> ({modId: val}, [unknown names])."""
    out: dict[int, int] = {}
    unknown: list[str] = []
    if not s:
        return out, unknown
    table = name_to_id()
    for raw_name, raw_val in _TOKEN_RE.findall(s):
        name = raw_name.strip().upper().replace(" ", "_")
        if name not in table:
            unknown.append(raw_name.strip())
            continue
        out[table[name]] = out.get(table[name], 0) + int(raw_val)
    return out, unknown


def _item_by_ref(conn, ref: str) -> tuple[int | None, dict[int, int]]:
    """Resolve a 'Base item' cell (id or name) to (item_id, base item mods)."""
    ref = ref.strip()
    if not ref:
        return None, {}
    if ref.isdigit():
        item_id = int(ref)
    else:
        cur = conn.cursor()
        needle = ref.lower().replace(" ", "_")
        cur.execute("SELECT itemId FROM item_equipment WHERE name=%s", (needle,))
        row = cur.fetchone()
        if not row:
            return None, {}
        item_id = row["itemId"]
    return item_id, gear.item_base_mods(conn, item_id)


def _slot_mods(conn, slots: dict[int, gear.EquippedSlot], slot_id: int) -> dict[int, int]:
    s = slots.get(slot_id)
    if not s:
        return {}
    return gear.item_total_mods(conn, s.item_id, s.extra)


@dataclass
class ScenarioResult:
    auto: float
    ws: float
    rotation: float
    note: str


def score_scenario(conn, anchor, equipped_mods, slots, wctx, ctx: WhatIfCtx,
                   slot_name: str, base_item: str, mod_str: str) -> ScenarioResult:
    """Build the candidate set for one scenario row and evaluate it exactly."""
    notes: list[str] = []
    slot_id = gear.SLOT_IDS.get(slot_name.strip().lower())
    if slot_id is None:
        return ScenarioResult(0, 0, 0, f"unknown slot {slot_name!r}")
    if slot_name.strip().lower() in ("main", "sub", "range", "ammo"):
        notes.append("weapon slot: anchor assumes a FIXED weapon; result approximate")

    item_id, item_base = _item_by_ref(conn, base_item)
    if base_item.strip() and item_id is None:
        notes.append(f"base item {base_item!r} not found; using mods only")
    add_mods, unknown = parse_mod_string(mod_str)
    if unknown:
        notes.append("unknown mods: " + ", ".join(unknown))

    # Candidate slot mods = base item's own mods + the typed (augment/extra) mods.
    new_slot: dict[int, int] = dict(item_base)
    for m, v in add_mods.items():
        new_slot[m] = new_slot.get(m, 0) + v

    # Candidate SET = equipped total - the piece we're replacing + the new piece.
    cand = dict(equipped_mods)
    for m, v in _slot_mods(conn, slots, slot_id).items():
        cand[m] = cand.get(m, 0) - v
    for m, v in new_slot.items():
        cand[m] = cand.get(m, 0) + v

    cs = _build(anchor, equipped_mods, cand, wctx)
    a, w, r = metric_triplet(cs, ctx)
    return ScenarioResult(a, w, r, "; ".join(notes))


# ---------------------------------------------------------------------------
# Workbook styling helpers
# ---------------------------------------------------------------------------
_HDR_FILL = PatternFill("solid", fgColor="1F4E78")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
_TITLE_FONT = Font(bold=True, size=14)
_SUB_FONT = Font(italic=True, color="555555")
_TOTAL_FONT = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _hdr(ws, row: int, headers: list[str]) -> None:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _widths(ws, widths: dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Workbook generation
# ---------------------------------------------------------------------------
def build_workbook(path: str, conn, charname: str, anchor, equipped_mods,
                   slots, wctx, jobs, ctx: WhatIfCtx) -> None:
    base, sens = compute_sensitivity(anchor, equipped_mods, wctx, ctx)
    base_cs = _build(anchor, equipped_mods, equipped_mods, wctx)

    wb = Workbook()
    _sheet_howto(wb, ctx)
    _sheet_baseline(wb, charname, anchor, ctx, base)
    _sheet_weaponskills(wb, conn, base_cs, wctx, jobs, ctx)
    _sheet_targets(wb, base_cs, ctx)
    _sheet_sensitivity(wb, sens)
    _sheet_scenarios(wb, base)
    _sheet_ctx(wb, charname, ctx)
    wb.save(path)


def _sheet_howto(wb: Workbook, ctx: WhatIfCtx) -> None:
    ws = wb.active
    ws.title = "How to use"
    _widths(ws, {"A": 100})
    lines = [
        ("DPS / Weaponskill what-if workbook", _TITLE_FONT),
        (f"Character: {ctx.charname}   Target: {ctx.target.label} "
         f"(T{ctx.target.tier})   WS column: "
         f"{_label(ctx.ws.name) if ctx.ws else '(none)'} @ TP {ctx.tp}", _SUB_FONT),
        ("", None),
        ("Pick what to analyze (auto-picked for you, but you choose):", _TOTAL_FONT),
        ("   * WEAPONSKILLS sheet -- every WS your weapon can use, ranked by "
         "damage on your current set. The auto-pick is marked.", None),
        ("   * TARGETS sheet -- your auto-DPS / WS / rotation vs all 15 Hunting "
         "League NMs. The current target is marked.", None),
        ("   To re-focus every sheet on a different WS or NM, re-run whatif with "
         '--ws "Name" and/or --target "Name".', None),
        ("", None),
        ("Two ways to test 'is this gear/augment worth it?':", _TOTAL_FONT),
        ("", None),
        ("1) SENSITIVITY sheet  (instant, approximate)", _TOTAL_FONT),
        ("   Each row is the value of ONE point of a mod at your CURRENT set, for "
         "auto-DPS, weaponskill, and rotation.", None),
        ("   Type how much of each mod a hypothetical augment/piece would give in "
         "the yellow 'Your amount' column;", None),
        ("   the green 'Est.' columns and the TOTAL row update live. A value near "
         "0 means you're capped there -- don't bother.", None),
        ("   This is a first-order estimate: great for quick triage, but big "
         "combined swings drift from exact.", None),
        ("", None),
        ("2) SCENARIOS sheet  (exact, needs one re-run)", _TOTAL_FONT),
        ("   Fill a row per hypothetical: a name, the Slot, an optional Base item "
         "(owned or aspirational), and a Mods list.", None),
        ("   Mods syntax:  STR:15 ATT:20 DOUBLE_ATTACK:3   (also accepts + or = "
         "or spaces; names are the mod.lua enum names).", None),
        ("   Base item blank = a brand-new item defined entirely by your Mods "
         "list. Base item set = that item's stats PLUS your Mods (augments).", None),
        ("   Then run:   python -m tools.dpscalc score "
         f"{ctx.charname} <your_mystats.txt> <this_file.xlsx>", None),
        ("   The exact Auto/WS/Rotation columns and the deltas vs your equipped "
         "set get written back into the sheet.", None),
        ("", None),
        ("Notes", _TOTAL_FONT),
        ("   * The weapon (main+sub) is held FIXED. A weapon swap needs a fresh "
         "!mystats anchor captured with that weapon equipped.", None),
        ("   * Capture !mystats UNBUFFED and WITHOUT FOOD in your currently "
         "equipped set, or the numbers drift.", None),
        ("   * Mod names: run  python -m tools.dpscalc mods  for a sample, or see "
         "scripts/enum/mod.lua for the full list.", None),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = _WRAP
        if font:
            cell.font = font


def _sheet_baseline(wb: Workbook, charname: str, anchor, ctx: WhatIfCtx,
                    base: tuple[float, float, float]) -> None:
    ws = wb.create_sheet("Baseline")
    _widths(ws, {"A": 26, "B": 18})
    rows = [
        ("Equipped-set baseline", None),
        (f"{charname}  {anchor.main_job}{anchor.main_lvl}/"
         f"{anchor.sub_job}{anchor.sub_lvl}", _SUB_FONT),
        ("", None),
        ("Target", ctx.target.label),
        ("Target DEF", ctx.target.defense()),
        ("Target EVA", ctx.target.evasion()),
        ("", None),
        ("Auto-attack DPS", round(base[0], 1)),
        (f"WS: {_label(ctx.ws.name) if ctx.ws else '(none)'} @ TP {ctx.tp}",
         round(base[1], 0)),
        ("Rotation DPS", round(base[2], 1)),
        ("", None),
        ("STR", anchor.stats.get("STR", 0)),
        ("DEX", anchor.stats.get("DEX", 0)),
        ("Attack", anchor.att),
        ("Accuracy", anchor.acc),
        ("Weapon DMG", anchor.wpn_dmg),
        ("Dbl/Trpl/Quad Atk %", f"{anchor.da}/{anchor.ta}/{anchor.qa}"),
        ("Dual Wield %", anchor.dual_wield),
        ("Haste-Gear (/1024)", anchor.haste_gear),
    ]
    ws.cell(row=1, column=1).font = _TITLE_FONT
    for i, (k, v) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=k)
        if v is not None and not isinstance(v, PatternFill) and i > 2:
            ws.cell(row=i, column=2, value=v)
        if k in ("Auto-attack DPS", "Rotation DPS") or k.startswith("WS:"):
            ws.cell(row=i, column=1).font = _TOTAL_FONT
            ws.cell(row=i, column=2).font = _TOTAL_FONT
    ws.cell(row=2, column=1).font = _SUB_FONT


def _sheet_weaponskills(wb: Workbook, conn, base_cs, wctx, jobs,
                        ctx: WhatIfCtx) -> None:
    """Menu of every WS usable with the equipped weapon, ranked by damage vs the
    chosen target. The auto-picked WS is marked; re-run with --ws to switch."""
    ws = wb.create_sheet("Weaponskills")
    _widths(ws, {"A": 22, "B": 6, "C": 24, "D": 12, "E": 12, "F": 12, "G": 16})
    ws.cell(row=1, column=1,
            value=f"Weaponskills vs {ctx.target.label} -- ranked by your "
                  f"configured TP{ctx.tp} damage on your CURRENT set "
                  f"(columns show the full TP curve)").font = _TITLE_FONT
    _hdr(ws, 2, ["Weaponskill", "Hits", "WSC (stat blend)",
                 "dmg @ TP1000", "dmg @ TP2000", "dmg @ TP3000", "Pick"])
    ws.freeze_panes = "A3"

    ws_list = wsdata.ws_for_weapon(conn, wctx.main_skill,
                                   jobs.get("mjob"), jobs.get("sjob"))
    rows = []
    for w in ws_list:
        wsc = ", ".join(f"{int(v * 100)}% {k}"
                        for k, v in sorted(w.wsc.items(), key=lambda kv: -kv[1]))
        d = {tp: engine.ws_damage(base_cs, ctx.target, w, tp).damage
             for tp in (1000, 2000, 3000)}
        # Rank by the workbook's configured TP so the auto-pick (chosen at the
        # same TP) sorts to the top; fall back to the dict if tp is 1k/2k/3k.
        d_rank = d.get(ctx.tp) or engine.ws_damage(
            base_cs, ctx.target, w, ctx.tp).damage
        rows.append((w.name, w.num_hits, wsc, d, d_rank))
    rows.sort(key=lambda r: r[4], reverse=True)

    picked = ctx.ws.name if ctx.ws else None
    for i, (name, nhits, wsc, d, _d_rank) in enumerate(rows):
        row = 3 + i
        ws.cell(row=row, column=1, value=_label(name))
        ws.cell(row=row, column=2, value=nhits)
        ws.cell(row=row, column=3, value=wsc or "-")
        for c, tp in ((4, 1000), (5, 2000), (6, 3000)):
            ws.cell(row=row, column=c, value=round(d[tp])).number_format = "0"
        if name == picked:
            ws.cell(row=row, column=7, value=f"* auto-pick @ TP{ctx.tp}")
            for c in range(1, 8):
                ws.cell(row=row, column=c).font = _TOTAL_FONT
    note = 3 + len(rows) + 1
    ws.cell(row=note, column=1, value=(
        'Want the workbook focused on a different WS? Re-run: '
        'whatif ... --ws "Savage Blade"  (Baseline / Sensitivity / Scenarios '
        'all follow the chosen WS).')).alignment = _WRAP
    ws.merge_cells(start_row=note, start_column=1, end_row=note, end_column=7)
    ws.cell(row=note, column=1).font = _SUB_FONT


def _sheet_targets(wb: Workbook, base_cs, ctx: WhatIfCtx) -> None:
    """Menu of the char's auto-DPS / WS / rotation vs every Hunting League NM,
    so you can see which target to analyze. Current target is marked."""
    ws = wb.create_sheet("Targets")
    _widths(ws, {"A": 24, "B": 6, "C": 8, "D": 8, "E": 6, "F": 12, "G": 14,
                 "H": 14})
    wsname = _label(ctx.ws.name) if ctx.ws else "(none)"
    ws.cell(row=1, column=1,
            value=f"Your set vs every Hunting League NM  "
                  f"(WS = {wsname} @ TP{ctx.tp})").font = _TITLE_FONT
    _hdr(ws, 2, ["Target", "Tier", "DEF", "EVA", "Pts", "Auto DPS",
                 "WS dmg", "Rotation DPS"])
    ws.freeze_panes = "A3"

    tlist = sorted(targets.load_targets(),
                   key=lambda t: (t.tier, t.points, t.label))
    cur = ctx.target.name
    for i, t in enumerate(tlist):
        row = 3 + i
        auto = engine.auto_attack_dps(base_cs, t).dps
        if ctx.ws:
            wsd = engine.ws_damage(base_cs, t, ctx.ws, ctx.tp).damage
            rot = engine.rotation_dps(base_cs, t, ctx.ws, ctx.tp).total_dps
        else:
            wsd, rot = 0.0, auto
        marked = t.name == cur
        ws.cell(row=row, column=1, value=t.label + ("  *" if marked else ""))
        ws.cell(row=row, column=2, value=t.tier)
        ws.cell(row=row, column=3, value=t.defense())
        ws.cell(row=row, column=4, value=t.evasion())
        ws.cell(row=row, column=5, value=t.points)
        ws.cell(row=row, column=6, value=round(auto, 1)).number_format = "0.0"
        ws.cell(row=row, column=7, value=round(wsd)).number_format = "0"
        ws.cell(row=row, column=8, value=round(rot, 1)).number_format = "0.0"
        if marked:
            for c in range(1, 9):
                ws.cell(row=row, column=c).font = _TOTAL_FONT
    note = 3 + len(tlist) + 1
    ws.cell(row=note, column=1, value=(
        'Want the workbook focused on a different NM? Re-run: '
        'whatif ... --target "Kirin"  (the * row is the one all other sheets '
        'use).')).alignment = _WRAP
    ws.merge_cells(start_row=note, start_column=1, end_row=note, end_column=8)
    ws.cell(row=note, column=1).font = _SUB_FONT


def _sheet_sensitivity(wb: Workbook, sens) -> None:
    ws = wb.create_sheet("Sensitivity")
    _widths(ws, {"A": 22, "B": 16, "C": 14, "D": 14, "E": 14, "F": 13,
                 "G": 14, "H": 14, "I": 14})
    ws.cell(row=1, column=1, value="Sensitivity -- value of +1 of each mod "
            "(type amounts in the yellow column)").font = _TITLE_FONT
    headers = ["Mod", "Unit", "+1 Auto DPS", "+1 WS dmg", "+1 Rotation",
               "Your amount", "Est. Auto", "Est. WS", "Est. Rotation"]
    _hdr(ws, 2, headers)
    ws.freeze_panes = "A3"

    first = 3
    for i, (label, unit, a, w, r) in enumerate(sens):
        row = first + i
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=unit)
        ws.cell(row=row, column=3, value=round(a, 4)).number_format = "0.000"
        ws.cell(row=row, column=4, value=round(w, 3)).number_format = "0.00"
        ws.cell(row=row, column=5, value=round(r, 4)).number_format = "0.000"
        amt = ws.cell(row=row, column=6, value=0)
        amt.fill = _INPUT_FILL
        amt.number_format = "0.##"
        ws.cell(row=row, column=7,
                value=f"=C{row}*F{row}").number_format = "0.0"
        ws.cell(row=row, column=8,
                value=f"=D{row}*F{row}").number_format = "0"
        ws.cell(row=row, column=9,
                value=f"=E{row}*F{row}").number_format = "0.0"

    last = first + len(sens) - 1
    trow = last + 1
    ws.cell(row=trow, column=1, value="TOTAL (typed package)").font = _TOTAL_FONT
    for col in (7, 8, 9):
        L = get_column_letter(col)
        c = ws.cell(row=trow, column=col, value=f"=SUM({L}{first}:{L}{last})")
        c.font = _TOTAL_FONT
        c.number_format = "0.0" if col != 8 else "0"

    note = trow + 2
    ws.cell(row=note, column=1, value=(
        "Estimate = your amount x per-point value, summed. First-order: exact "
        "for linear mods, ~0 when you're capped, drifts on big combined swings. "
        "Use the Scenarios sheet for exact numbers."))
    ws.cell(row=note, column=1).alignment = _WRAP
    ws.merge_cells(start_row=note, start_column=1, end_row=note, end_column=9)


_SCN_HEADERS = ["Scenario", "Slot", "Base item (optional)",
                "Mods (e.g. STR:15 ATT:20 DOUBLE_ATTACK:3)",
                "Auto DPS", "WS dmg", "Rotation DPS",
                "dAuto", "dWS", "dRotation", "Notes"]
# Example rows demonstrate both modelling styles; the user overwrites them.
_SCN_EXAMPLES = [
    ["(equipped baseline)", "", "", "", None, None, None, 0, 0, 0,
     "filled by `score`"],
    ["ex: augment an earring", "ear1", "brutal_earring",
     "STR:8 DOUBLE_ATTACK:3", None, None, None, None, None, None,
     "base item + extra augment mods"],
    ["ex: brand-new ring", "ring1", "", "STR:10 ATT:15 CRITHITRATE:4",
     None, None, None, None, None, None, "mods only (no base item)"],
]


def _sheet_scenarios(wb: Workbook, base: tuple[float, float, float]) -> None:
    ws = wb.create_sheet("Scenarios")
    _widths(ws, {"A": 24, "B": 8, "C": 22, "D": 42, "E": 11, "F": 11,
                 "G": 12, "H": 10, "I": 10, "J": 11, "K": 34})
    ws.cell(row=1, column=1, value="Scenarios -- fill rows, then run `score` "
            "to write exact numbers").font = _TITLE_FONT
    _hdr(ws, 2, _SCN_HEADERS)
    ws.freeze_panes = "E3"

    for i, ex in enumerate(_SCN_EXAMPLES):
        row = 3 + i
        for c, val in enumerate(ex, start=1):
            if val is None:
                continue
            ws.cell(row=row, column=c, value=val)
        for c in range(1, 5):                      # input cells -> yellow
            ws.cell(row=row, column=c).fill = _INPUT_FILL
    # Baseline reference numbers on the first example row.
    ws.cell(row=3, column=5, value=round(base[0], 1))
    ws.cell(row=3, column=6, value=round(base[1], 0))
    ws.cell(row=3, column=7, value=round(base[2], 1))

    # Blank yellow input rows for the user to fill.
    for row in range(6, 26):
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = _INPUT_FILL


def _sheet_ctx(wb: Workbook, charname: str, ctx: WhatIfCtx) -> None:
    ws = wb.create_sheet("_ctx")
    pairs = {
        "charname": charname,
        "target": ctx.target.name,
        "ws": ctx.ws.name if ctx.ws else "",
        "tp": ctx.tp,
        "main_skill": ctx.main_skill,
    }
    for i, (k, v) in enumerate(pairs.items(), start=1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    ws.sheet_state = "hidden"


def read_ctx(path: str) -> dict[str, str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "_ctx" not in wb.sheetnames:
        return {}
    ws = wb["_ctx"]
    out = {}
    for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
        if row and row[0] is not None:
            out[str(row[0])] = row[1]
    return out


# ---------------------------------------------------------------------------
# Scenario scoring (write results back into the workbook)
# ---------------------------------------------------------------------------
def score_workbook(path: str, conn, anchor, equipped_mods, slots, wctx,
                   ctx: WhatIfCtx) -> int:
    """Evaluate every filled Scenarios row through the engine; write results
    back in place. Returns the number of scenarios scored."""
    wb = load_workbook(path)            # keep formulas (Sensitivity sheet) intact
    if "Scenarios" not in wb.sheetnames:
        raise ValueError("workbook has no 'Scenarios' sheet")
    ws = wb["Scenarios"]
    base = metric_triplet(_build(anchor, equipped_mods, equipped_mods, wctx), ctx)

    scored = 0
    for row in range(3, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        slot = ws.cell(row=row, column=2).value
        base_item = ws.cell(row=row, column=3).value
        mods = ws.cell(row=row, column=4).value
        # The reference row carries no slot/mods -- just refresh its baseline.
        if name and not slot and not mods:
            ws.cell(row=row, column=5, value=round(base[0], 1))
            ws.cell(row=row, column=6, value=round(base[1], 0))
            ws.cell(row=row, column=7, value=round(base[2], 1))
            ws.cell(row=row, column=8, value=0)
            ws.cell(row=row, column=9, value=0)
            ws.cell(row=row, column=10, value=0)
            continue
        if not slot and not mods:
            continue
        res = score_scenario(conn, anchor, equipped_mods, slots, wctx, ctx,
                             str(slot or ""), str(base_item or ""), str(mods or ""))
        ws.cell(row=row, column=5, value=round(res.auto, 1)).number_format = "0.0"
        ws.cell(row=row, column=6, value=round(res.ws, 0)).number_format = "0"
        ws.cell(row=row, column=7, value=round(res.rotation, 1)).number_format = "0.0"
        ws.cell(row=row, column=8, value=round(res.auto - base[0], 1)).number_format = "+0.0;-0.0"
        ws.cell(row=row, column=9, value=round(res.ws - base[1], 0)).number_format = "+0;-0"
        ws.cell(row=row, column=10, value=round(res.rotation - base[2], 1)).number_format = "+0.0;-0.0"
        ws.cell(row=row, column=11, value=res.note)
        scored += 1
    wb.save(path)
    return scored


def _label(name: str) -> str:
    return name.replace("_", " ").title()


# ---------------------------------------------------------------------------
# Default-WS selection: the highest-damage WS for the weapon at a reference TP.
# ---------------------------------------------------------------------------
def pick_default_ws(conn, base_cs: state.CombatStats, wctx, jobs, tgt,
                    tp_ref: int = 2000):
    """Return (all usable WS for this weapon/job, the highest-damage one)."""
    ws_list = wsdata.ws_for_weapon(conn, wctx.main_skill,
                                   jobs.get("mjob"), jobs.get("sjob"))
    if not ws_list:
        return ws_list, None
    best = max(ws_list,
               key=lambda w: engine.ws_damage(base_cs, tgt, w, tp_ref).damage)
    return ws_list, best
