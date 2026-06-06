"""Self-contained Excel DPS workbook generator.

Unlike whatif.py (which leaves the heavy math in Python and only writes a
sensitivity grid), this module bakes the ENTIRE combat model into Excel cell
formulas and ships every curated gear piece + augment from the live server as
selectable dropdowns. The result is a single .xlsx the player opens and drives
with no Python in the loop: pick an item / augment per slot, pick a target +
weaponskill + TP, and the auto-attack DPS, weaponskill damage and rotation DPS
recompute live.

How the math stays correct without re-running the engine
--------------------------------------------------------
The workbook uses the same anchored-delta model as state.py. `!mystats` gives the
FINAL stats of the equipped set (the anchor, an editable input on the Stats
sheet). The weapon is FIXED, so swapping the 14 armor/accessory slots only moves
gear mods. For each tracked mod the workbook computes
    delta = (candidate 14-slot total) - (equipped 14-slot total)
and rebuilds every final number the engine needs:
  * STR..CHR, crit, multi-attack, haste, DW, Store TP, WS dmg  = anchor + delta
  * ATT  = invert the anchor's ATTP multiplier to a core, swap mod[ATT] +
           floor(STR*strMult), re-apply ATTP  (same as state._core_from_final_att)
  * ACC  = anchor + delta(ACC) + floor(DEX*dexMult) re-base
  * gear fTP (ANY_FTP_BONUS/256) and WSACC are gear-only, so they use the full
    candidate gear total (weapon baked in + 14-slot total)
then ports formulas.py + engine.py onto a Calc sheet (fSTR, the pDIF expected
value, hit rate, swing crit, multi-attack EV, WSC, fTP, TP return, swing
interval, auto DPS, WS damage, rotation).

Scope: single-wield only (the deliverable target, Jbae PLD/NIN, wields a sword +
grip). A dual-wield / H2H build swings the offhand too; that path raises so the
caller can't silently get half a model.
"""
from __future__ import annotations

import math
import re
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

from . import db, engine, gear, state, targets, wsdata
from .anchor import Anchor
from .formulas import PDIF_WEAPON_CAP, tp_return_pc, weapon_rank
from .gameconf import CONF
from .mods import name_to_id

# ---------------------------------------------------------------------------
# Tracked mods (the gear-swap columns) and the offensive-mod curation filter.
# ---------------------------------------------------------------------------
TRACKED = [
    "STR", "DEX", "VIT", "AGI", "INT", "MND", "CHR",
    "ATT", "ATTP", "ACC",
    "CRITHITRATE", "CRIT_DMG_INCREASE",
    "DOUBLE_ATTACK", "TRIPLE_ATTACK", "QUAD_ATTACK",
    "HASTE_GEAR", "HASTE_MAGIC", "HASTE_ABILITY",
    "DUAL_WIELD", "STORETP", "TP_BONUS",
    "ALL_WSDMG_ALL_HITS", "ALL_WSDMG_FIRST_HIT",
    "ANY_FTP_BONUS", "WSACC",
]
# An item earns a spot in the curated list if it carries >=1 of these.
OFFENSIVE = {
    "ATT", "ATTP", "ACC", "STR", "DEX",
    "DOUBLE_ATTACK", "TRIPLE_ATTACK", "QUAD_ATTACK",
    "CRITHITRATE", "CRIT_DMG_INCREASE", "STORETP", "DUAL_WIELD",
    "HASTE_GEAR", "ALL_WSDMG_ALL_HITS", "ALL_WSDMG_FIRST_HIT",
    "WSACC", "ANY_FTP_BONUS", "TP_BONUS",
}

# The 14 swap slots (everything except main/sub weapon): range..back.
SWAP_SLOTS = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]

_MIDS = {n: name_to_id()[n] for n in TRACKED}

# Build-sheet cell schema (fixed up front so every sheet can reference it).
B_TARGET = "Build!$B$4"
B_WS = "Build!$B$5"
B_TP = "Build!$B$6"
SLOT_ROW0 = 14            # first per-slot row on the Build sheet
# per-slot columns: B item, C augment, D count, E free1 name, F free1 val,
#                   G free2 name, H free2 val
OUT_AUTO = "Build!$B$9"
OUT_WS = "Build!$B$10"
OUT_ROT = "Build!$B$11"

HDR_FILL = PatternFill("solid", fgColor="2F5496")
HDR_FONT = Font(bold=True, color="FFFFFF")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
IN_FILL = PatternFill("solid", fgColor="FFF2CC")    # editable input cells
OUT_FILL = PatternFill("solid", fgColor="E2EFDA")   # live output cells
BOLD = Font(bold=True)


# ---------------------------------------------------------------------------
# Data export
# ---------------------------------------------------------------------------
@dataclass
class GearRow:
    item_id: int
    name: str           # display name (unique)
    slot_mask: int
    mods: dict[str, int]  # tracked mod -> value


def _proj(mod_by_id: dict[int, int], cols=TRACKED, mids=_MIDS) -> dict[str, int]:
    return {n: mod_by_id.get(mids[n], 0) for n in cols}


def curate_gear(conn, jobmask: int) -> list[GearRow]:
    """Level>=99 (or ilevel>=99), equippable by the job pair, fits a swap slot,
    and carries at least one offensive mod. Names are made globally unique."""
    cur = conn.cursor()
    cur.execute(
        "SELECT itemId, name, slot, jobs, level, ilevel FROM item_equipment "
        "WHERE (level >= 99 OR ilevel >= 99) AND (jobs & %s) <> 0",
        (jobmask,),
    )
    out: list[GearRow] = []
    used: dict[str, int] = {}
    for e in cur.fetchall():
        slot_mask = e["slot"]
        if not any(slot_mask & (1 << s) for s in SWAP_SLOTS):
            continue
        base = gear.item_base_mods(conn, e["itemId"])
        proj = _proj(base)
        if not any(proj[m] for m in OFFENSIVE):
            continue
        name = e["name"]
        if name in used:                       # disambiguate duplicate names
            name = f"{name} ({e['itemId']})"
        used[name] = e["itemId"]
        out.append(GearRow(e["itemId"], name, slot_mask, proj))
    out.sort(key=lambda r: r.name.lower())
    return out


@dataclass
class AugRow:
    label: str
    aug_id: int
    mods: dict[str, int]


_AUG_ENTRY = re.compile(r"\{([^}]*)\}")
_AUG_ID = re.compile(r"augId\s*=\s*(\d+)")
_AUG_LABEL = re.compile(r"label\s*=\s*'([^']*)'")


def _augment_raw_by_id(conn, aug_id: int) -> dict[int, int]:
    """Per-single-catalyst magnitudes (modId -> value) for one augId at exdata=0.

    Mirrors gear.augment_mods with value(exdata)=0: magnitude per augments-table
    row = row.value * (multiplier>1 ? multiplier : 1), summed by modId. N stacked
    catalysts deliver N slots, so the workbook multiplies this by the count."""
    by_id: dict[int, int] = {}
    for row in gear._augment_table(conn).get(aug_id, []):
        if row["isPet"]:
            continue
        mod_id = row["modId"]
        if mod_id == 0:
            continue
        mult = row["multiplier"]
        mag = row["value"] * (mult if mult > 1 else 1)
        by_id[mod_id] = by_id.get(mod_id, 0) + mag
    return by_id


def _augment_tracked_mods(conn, aug_id: int) -> dict[str, int]:
    return _proj(_augment_raw_by_id(conn, aug_id))


def build_augdb(conn) -> list[AugRow]:
    """Resolve the Augment Moogle catalog to offensive augments (touching >=1
    tracked mod), deduped by augId, labels made unique for MATCH."""
    path = db.REPO_ROOT / "modules" / "custom" / "lua" / "augment_catalog.lua"
    text = path.read_text(encoding="utf-8", errors="replace")
    seen: dict[int, str] = {}
    rows: list[AugRow] = []
    used_labels: dict[str, int] = {}
    for m in _AUG_ENTRY.finditer(text):
        body = m.group(1)
        am, lm = _AUG_ID.search(body), _AUG_LABEL.search(body)
        if not am or not lm:
            continue
        aug_id = int(am.group(1))
        if aug_id in seen:
            continue
        proj = _augment_tracked_mods(conn, aug_id)
        if not any(proj.values()):
            continue
        seen[aug_id] = lm.group(1)
        label = lm.group(1)
        if label in used_labels:
            label = f"{label} #{aug_id}"
        used_labels[label] = aug_id
        rows.append(AugRow(label, aug_id, proj))
    rows.sort(key=lambda r: r.label.lower())
    return rows


def build_wsdb(conn, wctx, mjob: int, sjob: int):
    """Physical WS usable with the fixed weapon, each flattened to the row schema
    the Calc sheet INDEXes (ftp/atk/ign/crit tables expanded to 3 cells each)."""
    ws_list = wsdata.ws_for_weapon(conn, wctx.main_skill, mjob, sjob)
    ws_list.sort(key=lambda w: w.name)
    return ws_list


# ---------------------------------------------------------------------------
# Calc sheet helper
# ---------------------------------------------------------------------------
class Calc:
    """Append-only scalar machinery on one sheet; tracks logical name -> $B$row."""

    def __init__(self, ws):
        self.ws = ws
        self.row = 1
        self.ref: dict[str, str] = {}
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 60

    def section(self, title: str) -> None:
        c = self.ws.cell(self.row, 1, title)
        c.font = BOLD
        c.fill = SUB_FILL
        self.ws.cell(self.row, 2).fill = SUB_FILL
        self.row += 1

    def put(self, name: str | None, label: str, value, note: str = "") -> str:
        r = self.row
        self.ws.cell(r, 1, label)
        self.ws.cell(r, 2).value = value
        if note:
            self.ws.cell(r, 3, note)
        ref = f"Calc!$B${r}"
        if name:
            self.ref[name] = ref
        self.row += 1
        return ref


def _hit_rate(acc_expr: str, eva_ref: str, cap: float) -> str:
    return f"MAX(0.2,MIN({cap},(75+(({acc_expr})-{eva_ref})/2)/100))"


def _ftp(tp_ref: str, f1: str, f2: str, f3: str) -> str:
    return (f"IF({tp_ref}<1000,1,IF({tp_ref}>=2000,"
            f"{f2}+({tp_ref}-2000)*({f3}-{f2})/1000,"
            f"{f1}+({tp_ref}-1000)*({f2}-{f1})/1000))")


def _emit_pdif(calc: Calc, prefix: str, attack_ref: str, atk_mult: str,
               ignore: str, crit_rate_ref: str, *, tgt_def: str,
               crit_dmg_mult_ref: str, cap: float) -> str:
    """Lay down the melee_pdif_ev expected-value block; return the pDIF cell."""
    actor = calc.put(None, f"{prefix} actor ATT",
                     f"=MAX(1,INT({attack_ref}*({atk_mult})))")
    deff = calc.put(None, f"{prefix} eff DEF",
                    f"=MAX(1,INT({tgt_def}*(1-({ignore}))))")
    ratio = calc.put(None, f"{prefix} base ratio", f"=({actor})/({deff})")

    def branch(tag: str, crit: bool) -> str:
        w = calc.put(None, f"{prefix} {tag} wRatio",
                     f"=({ratio})+{1 if crit else 0}")
        cp = calc.put(None, f"{prefix} {tag} cap",
                      f"={cap}+{1 if crit else 0}")
        spike = calc.put(None, f"{prefix} {tag} spike",
                         f"=IF(AND({w}>0.5,{w}<1.5),"
                         f"MAX(0,MIN(1/3,(0.5-ABS({w}-1))*1.2)),0)")
        up = calc.put(None, f"{prefix} {tag} up",
                      f"=IF({w}<0.5,{w}+0.5,IF({w}<0.7,1,IF({w}<1.2,{w}+0.3,"
                      f"IF({w}<1.5,{w}*1.25,MIN({w}+0.375,{cp})))))")
        lo = calc.put(None, f"{prefix} {tag} lo",
                      f"=MAX(0,IF({w}<0.38,0,IF({w}<1.25,{w}*1176/1024-448/1024,"
                      f"IF({w}<1.51,1,IF({w}<2.44,{w}*1176/1024-775/1024,"
                      f"MIN({w}-0.375,{cp}))))))")
        ev = calc.put(None, f"{prefix} {tag} ev", f"=(({lo})+({up}))/2*1.025")
        if crit:
            return calc.put(None, f"{prefix} {tag} branch",
                            f"=({spike})*1+(1-({spike}))*({ev})*({crit_dmg_mult_ref})")
        return calc.put(None, f"{prefix} {tag} branch",
                        f"=({spike})*1+(1-({spike}))*({ev})")

    non = branch("non", False)
    cri = branch("crit", True)
    return calc.put(f"{prefix}_pdif", f"{prefix} pDIF EV",
                    f"=(1-({crit_rate_ref}))*({non})+({crit_rate_ref})*({cri})")


# ---------------------------------------------------------------------------
# Workbook assembly
# ---------------------------------------------------------------------------
def build_book(conn, charid: int, anchor: Anchor, out_path: str) -> dict:
    slots = gear.equipped_items(conn, charid)
    if gear.SLOT_IDS["main"] not in slots:
        raise ValueError("char has no main weapon equipped")
    wctx = state.resolve_weapon_ctx(conn, charid, slots)
    if wctx.has_offhand:
        raise NotImplementedError(
            "buildbook currently models single-wield only; this char has an "
            "offhand weapon (dual-wield/H2H). The offhand swing isn't ported yet.")

    cur = conn.cursor()
    cur.execute("SELECT mjob, sjob FROM char_stats WHERE charid=%s", (charid,))
    jr = cur.fetchone() or {}
    mjob, sjob = jr.get("mjob") or 0, jr.get("sjob") or 0
    jobmask = ((1 << (mjob - 1)) if mjob else 0) | ((1 << (sjob - 1)) if sjob else 0)

    geardb = curate_gear(conn, jobmask)
    augdb = build_augdb(conn)
    ws_list = build_wsdb(conn, wctx, mjob, sjob)
    tgts = sorted(targets.load_targets(), key=lambda t: (t.tier, t.points, t.label))

    # Per-slot equipped mods (with real augments) + the fixed weapon contribution.
    equip_proj = {
        s: _proj(gear.item_total_mods(conn, slots[s].item_id, slots[s].extra))
        if s in slots else {n: 0 for n in TRACKED}
        for s in SWAP_SLOTS
    }
    weapon_slots = {sid: slots[sid] for sid in (0, 1) if sid in slots}
    weapon_raw = gear.gear_set_mods(conn, weapon_slots)
    weapon_baked = _proj(weapon_raw)

    # ---- Per-WS WS-damage columns (WEAPONSKILL_DAMAGE_BASE + ws_id) ----
    # These per-weaponskill gear mods are NOT reported by !mystats, so (like gear
    # fTP and WSACC) they sit outside the anchor and use the FULL candidate gear
    # total. Only ws_ids some weapon/equipped/curated/augment piece actually
    # carries become columns, so the workbook stays lean.
    BASE_WSDMG = name_to_id()["WEAPONSKILL_DAMAGE_BASE"]
    equip_raw = {
        s: (gear.item_total_mods(conn, slots[s].item_id, slots[s].extra)
            if s in slots else {})
        for s in SWAP_SLOTS
    }

    def _wid_carried(wid: int) -> bool:
        mid = BASE_WSDMG + wid
        return bool(
            weapon_raw.get(mid)
            or any(er.get(mid) for er in equip_raw.values())
            or any(gear.item_base_mods(conn, g.item_id).get(mid) for g in geardb)
            or any(_augment_raw_by_id(conn, a.aug_id).get(mid) for a in augdb))

    wsdmg_ids = sorted(wid for wid in {w.ws_id for w in ws_list} if _wid_carried(wid))
    extra_mids = {f"WSDMG_{wid}": BASE_WSDMG + wid for wid in wsdmg_ids}
    cols = TRACKED + list(extra_mids)

    # Fold the EXTRA per-WS columns into every projection (all built as TRACKED).
    for col, mid in extra_mids.items():
        for g in geardb:
            g.mods[col] = gear.item_base_mods(conn, g.item_id).get(mid, 0)
        for a in augdb:
            a.mods[col] = _augment_raw_by_id(conn, a.aug_id).get(mid, 0)
        for s in SWAP_SLOTS:
            equip_proj[s][col] = equip_raw[s].get(mid, 0)
        weapon_baked[col] = weapon_raw.get(mid, 0)

    wb = Workbook()
    wb.remove(wb.active)
    ws_guide = wb.create_sheet("Guide")
    ws_build = wb.create_sheet("Build")
    ws_stats = wb.create_sheet("Stats")
    ws_gc = wb.create_sheet("GearCalc")
    ws_calc = wb.create_sheet("Calc")
    ws_eb = wb.create_sheet("EquipBase")
    ws_gdb = wb.create_sheet("GearDB")
    ws_adb = wb.create_sheet("AugDB")
    ws_wdb = wb.create_sheet("WSDB")
    ws_tgt = wb.create_sheet("Targets")
    ws_lists = wb.create_sheet("Lists")

    # ---- Stats sheet (editable anchor inputs) ----
    stat_ref = _write_stats(ws_stats, anchor)

    # ---- GearDB ----
    _write_geardb(ws_gdb, geardb, cols)
    gdb_last = 1 + len(geardb)
    _name(wb, "GearDB_names", f"GearDB!$A$2:$A${gdb_last}")
    for j, m in enumerate(cols):
        col = get_column_letter(4 + j)
        _name(wb, f"GearDB_{m}", f"GearDB!${col}$2:${col}${gdb_last}")

    # ---- AugDB ----
    _write_augdb(ws_adb, augdb, cols)
    adb_last = 1 + len(augdb)
    _name(wb, "AugDB_labels", f"AugDB!$A$2:$A${adb_last}")
    for j, m in enumerate(cols):
        col = get_column_letter(3 + j)
        _name(wb, f"AugDB_{m}", f"AugDB!${col}$2:${col}${adb_last}")

    # ---- WSDB + Targets ----
    ws_last = _write_wsdb(ws_wdb, ws_list)
    tgt_last = _write_targets(ws_tgt, tgts)
    _name(wb, "list_targets", f"Targets!$A$2:$A${tgt_last}")
    _name(wb, "list_ws", f"WSDB!$A$2:$A${ws_last}")

    # ---- EquipBase (baked) ----
    eb_total = _write_equipbase(ws_eb, equip_proj, cols)

    # ---- Lists (dropdown sources) ----
    list_names = _write_lists(ws_lists, geardb, augdb)
    for nm, rng in list_names.items():
        _name(wb, nm, rng)

    # ---- GearCalc (live per-slot mod grid) ----
    gc_total = _write_gearcalc(ws_gc, cols)

    # ---- Calc (the ported combat math) ----
    calc = Calc(ws_calc)
    _write_calc(calc, wctx, stat_ref, gc_total, eb_total, weapon_baked,
                ws_last, tgt_last, wsdmg_ids)

    # ---- Build (interactive sheet) ----
    base_cs = state.build_combat_stats(anchor, gear.gear_set_mods(conn, slots),
                                       gear.gear_set_mods(conn, slots), wctx)
    default_target, default_ws, eng_base = _engine_baseline(base_cs, tgts, ws_list)
    _write_build(ws_build, calc.ref, default_target, default_ws,
                 eng_base, anchor)

    # ---- Guide ----
    _write_guide(ws_guide, anchor, len(geardb), len(augdb), len(ws_list), len(tgts))

    wb.active = wb.sheetnames.index("Build")
    wb.save(out_path)
    return {
        "gear": len(geardb), "aug": len(augdb), "ws": len(ws_list),
        "targets": len(tgts), "default_target": default_target.label,
        "default_ws": default_ws.name if default_ws else None,
        "engine_baseline": eng_base,
    }


def _name(wb, name: str, ref: str) -> None:
    wb.defined_names[name] = DefinedName(name, attr_text=ref)


# ---------------------------------------------------------------------------
# Stats sheet
# ---------------------------------------------------------------------------
_STAT_FIELDS = [
    ("str_a", "STR", "stats.STR"), ("dex_a", "DEX", "stats.DEX"),
    ("vit_a", "VIT", "stats.VIT"), ("agi_a", "AGI", "stats.AGI"),
    ("int_a", "INT", "stats.INT"), ("mnd_a", "MND", "stats.MND"),
    ("chr_a", "CHR", "stats.CHR"),
    ("att_a", "Melee ATT", "att"), ("acc_a", "Melee ACC", "acc"),
    ("wpn_dmg_a", "Wpn dmg", "wpn_dmg"),
    ("crit_rate_a", "Crit hit rate %", "crit_rate"),
    ("crit_dmg_a", "Crit dmg +%", "crit_dmg"),
    ("da_a", "Dbl Atk %", "da"), ("ta_a", "Trpl Atk %", "ta"),
    ("qa_a", "Quad Atk %", "qa"),
    ("haste_gear_a", "Haste-Gear (1%=100)", "haste_gear"),
    ("haste_magic_a", "Haste-Magic (1%=100)", "haste_magic"),
    ("haste_ability_a", "Haste-Ability (1%=100)", "haste_ability"),
    ("dual_wield_a", "Dual Wield %", "dual_wield"),
    ("store_tp_a", "Store TP", "store_tp"),
    ("tp_bonus_a", "TP Bonus", "tp_bonus"),
    ("ws_dmg_all_a", "WS dmg all hits %", "ws_dmg_all"),
    ("ws_dmg_first_a", "WS dmg 1st hit %", "ws_dmg_first"),
]


def _anchor_val(a: Anchor, path: str):
    if path.startswith("stats."):
        return a.stats.get(path.split(".")[1], 0)
    return getattr(a, path)


def _write_stats(ws, a: Anchor) -> dict:
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 50
    ws.cell(1, 1, "Your !mystats anchor").font = HDR_FONT
    ws.cell(1, 1).fill = HDR_FILL
    ws.cell(1, 2).fill = HDR_FILL
    ws.cell(2, 1, "EDIT these to match YOUR /mystats (unbuffed, no food).")
    ws.cell(2, 1).font = Font(italic=True, color="C00000")
    ref = {}
    r = 4
    for name, label, path in _STAT_FIELDS:
        ws.cell(r, 1, label)
        c = ws.cell(r, 2, _anchor_val(a, path))
        c.fill = IN_FILL
        ref[name] = f"Stats!$B${r}"
        r += 1
    return ref


# ---------------------------------------------------------------------------
# GearDB / AugDB / EquipBase / WSDB / Targets data sheets
# ---------------------------------------------------------------------------
def _header(ws, headers: list[str]) -> None:
    for j, h in enumerate(headers, start=1):
        c = ws.cell(1, j, h)
        c.font = HDR_FONT
        c.fill = HDR_FILL


def _write_geardb(ws, geardb: list[GearRow], cols=TRACKED) -> None:
    _header(ws, ["name", "slotmask", "itemId"] + cols)
    ws.column_dimensions["A"].width = 32
    for i, g in enumerate(geardb, start=2):
        ws.cell(i, 1, g.name)
        ws.cell(i, 2, g.slot_mask)
        ws.cell(i, 3, g.item_id)
        for j, m in enumerate(cols):
            ws.cell(i, 4 + j, g.mods[m])


def _write_augdb(ws, augdb: list[AugRow], cols=TRACKED) -> None:
    _header(ws, ["label", "augId"] + cols)
    ws.column_dimensions["A"].width = 34
    for i, a in enumerate(augdb, start=2):
        ws.cell(i, 1, a.label)
        ws.cell(i, 2, a.aug_id)
        for j, m in enumerate(cols):
            ws.cell(i, 3 + j, a.mods[m])


def _write_equipbase(ws, equip_proj: dict[int, dict[str, int]], cols=TRACKED) -> dict:
    _header(ws, ["slot"] + cols)
    ws.column_dimensions["A"].width = 10
    for i, s in enumerate(SWAP_SLOTS, start=2):
        ws.cell(i, 1, gear.SLOT_NAMES[s])
        for j, m in enumerate(cols):
            ws.cell(i, 2 + j, equip_proj[s][m])
    total_row = 2 + len(SWAP_SLOTS)
    ws.cell(total_row, 1, "TOTAL").font = BOLD
    total_ref = {}
    for j, m in enumerate(cols):
        col = get_column_letter(2 + j)
        ws.cell(total_row, 2 + j,
                f"=SUM({col}2:{col}{total_row - 1})").font = BOLD
        total_ref[m] = f"EquipBase!${col}${total_row}"
    return total_ref


def _write_targets(ws, tgts) -> int:
    _header(ws, ["name", "DEF", "EVA", "VIT", "AGI", "CRIT_EVA", "CRIT_DEF",
                 "tier", "label"])
    ws.column_dimensions["A"].width = 30
    for i, t in enumerate(tgts, start=2):
        ws.cell(i, 1, t.label)        # label is the user-facing, unique name
        ws.cell(i, 2, t.defense())
        ws.cell(i, 3, t.evasion())
        ws.cell(i, 4, t.vit())
        ws.cell(i, 5, t.agi())
        ws.cell(i, 6, t.crit_eva())
        ws.cell(i, 7, t.crit_def_bonus())
        ws.cell(i, 8, t.tier)
        ws.cell(i, 9, t.label)
    return 1 + len(tgts)


# WSDB column layout (1-based): 1 name, 2 numHits, 3-5 ftp, 6-12 wsc(STR..CHR),
# 13-15 atk, 16-18 ign, 19 has_ign, 20-22 crit, 23 has_crit, 24 multiHitFtp,
# 25 wsId.
def _write_wsdb(ws, ws_list) -> int:
    _header(ws, ["name", "numHits", "ftp1", "ftp2", "ftp3",
                 "STR", "DEX", "VIT", "AGI", "INT", "MND", "CHR",
                 "atk1", "atk2", "atk3", "ign1", "ign2", "ign3", "has_ign",
                 "crit1", "crit2", "crit3", "has_crit", "multiHitFtp", "wsId"])
    ws.column_dimensions["A"].width = 22
    for i, w in enumerate(ws_list, start=2):
        ftp = (w.ftp + [1.0, 1.0, 1.0])[:3]
        atk = (w.atk_varies or [1.0, 1.0, 1.0])
        atk = (list(atk) + [1.0, 1.0, 1.0])[:3]
        ign = w.ignored_def or [0.0, 0.0, 0.0]
        ign = (list(ign) + [0.0, 0.0, 0.0])[:3]
        crt = w.crit_varies or [0.0, 0.0, 0.0]
        crt = (list(crt) + [0.0, 0.0, 0.0])[:3]
        vals = [w.name, w.num_hits, ftp[0], ftp[1], ftp[2]]
        vals += [w.wsc.get(s, 0.0) for s in ("STR", "DEX", "VIT", "AGI", "INT", "MND", "CHR")]
        vals += [atk[0], atk[1], atk[2]]
        vals += [ign[0], ign[1], ign[2], 1 if w.ignored_def else 0]
        vals += [crt[0], crt[1], crt[2], 1 if w.crit_varies else 0]
        vals += [1 if w.multi_hit_ftp else 0, w.ws_id]
        for j, v in enumerate(vals, start=1):
            ws.cell(i, j, v)
    return 1 + len(ws_list)


# ---------------------------------------------------------------------------
# Lists sheet (per-slot item lists + augment + free-mod dropdown sources)
# ---------------------------------------------------------------------------
def _write_lists(ws, geardb: list[GearRow], augdb: list[AugRow]) -> dict[str, str]:
    names: dict[str, str] = {}
    col = 1
    for s in SWAP_SLOTS:
        letter = get_column_letter(col)
        ws.cell(1, col, gear.SLOT_NAMES[s]).font = BOLD
        entries = ["(keep equipped)", "(none)"]
        entries += [g.name for g in geardb if g.slot_mask & (1 << s)]
        for r, e in enumerate(entries, start=2):
            ws.cell(r, col, e)
        names[f"list_slot{s}"] = f"Lists!${letter}$2:${letter}${1 + len(entries)}"
        col += 1
    # augment list
    letter = get_column_letter(col)
    ws.cell(1, col, "augments").font = BOLD
    aug_entries = ["(none)"] + [a.label for a in augdb]
    for r, e in enumerate(aug_entries, start=2):
        ws.cell(r, col, e)
    names["list_augs"] = f"Lists!${letter}$2:${letter}${1 + len(aug_entries)}"
    col += 1
    # free-mod name list
    letter = get_column_letter(col)
    ws.cell(1, col, "freemods").font = BOLD
    free_entries = ["(none)"] + TRACKED
    for r, e in enumerate(free_entries, start=2):
        ws.cell(r, col, e)
    names["list_freemods"] = f"Lists!${letter}$2:${letter}${1 + len(free_entries)}"
    return names


# ---------------------------------------------------------------------------
# GearCalc sheet: 14 slot rows x 25 mod cols, + itemRow/augRow helpers + TOTAL.
# ---------------------------------------------------------------------------
def _write_gearcalc(ws, cols=TRACKED) -> dict:
    _header(ws, ["slot", "itemRow", "augRow"] + cols)
    ws.column_dimensions["A"].width = 10
    for i, s in enumerate(SWAP_SLOTS):
        r = 2 + i
        br = SLOT_ROW0 + i
        b_item = f"Build!$B${br}"
        b_aug = f"Build!$C${br}"
        b_cnt = f"Build!$D${br}"
        b_f1n, b_f1v = f"Build!$E${br}", f"Build!$F${br}"
        b_f2n, b_f2v = f"Build!$G${br}", f"Build!$H${br}"
        ws.cell(r, 1, gear.SLOT_NAMES[s])
        ws.cell(r, 2, f"=IFERROR(MATCH({b_item},GearDB_names,0),0)")
        ws.cell(r, 3, f"=IFERROR(MATCH({b_aug},AugDB_labels,0),0)")
        item_row = f"$B${r}"
        aug_row = f"$C${r}"
        for j, m in enumerate(cols):
            eb_cell = f"EquipBase!${get_column_letter(2 + j)}${r}"
            f = (
                f"=IF({b_item}=\"(keep equipped)\",{eb_cell},"
                f"IF(OR({b_item}=\"(none)\",{b_item}=\"\"),0,"
                f"IFERROR(INDEX(GearDB_{m},{item_row}),0)))"
                f"+IF({b_item}=\"(keep equipped)\",0,"
                f"IF({aug_row}>0,{b_cnt}*INDEX(AugDB_{m},{aug_row}),0)"
                f"+IF({b_f1n}=\"{m}\",{b_f1v},0)+IF({b_f2n}=\"{m}\",{b_f2v},0))"
            )
            ws.cell(r, 4 + j, f)
    total_row = 2 + len(SWAP_SLOTS)
    ws.cell(total_row, 1, "TOTAL").font = BOLD
    total_ref = {}
    for j, m in enumerate(cols):
        col = get_column_letter(4 + j)
        ws.cell(total_row, 4 + j,
                f"=SUM({col}2:{col}{total_row - 1})").font = BOLD
        total_ref[m] = f"GearCalc!${col}${total_row}"
    return total_ref


# ---------------------------------------------------------------------------
# Calc sheet: the ported combat model (single-wield).
# ---------------------------------------------------------------------------
def _write_calc(calc: Calc, wctx, stat, gc_total, eb_total, weapon_baked,
                ws_last, tgt_last, wsdmg_ids=()) -> None:
    R = calc.ref
    smul = wctx.str_mult_main
    dmul = wctx.dex_mult_main
    cap_main = PDIF_WEAPON_CAP.get(wctx.main_skill, 3.25)
    main_cap = 0.95 if wctx.is_2h else 0.99
    wrank = weapon_rank(
        wctx.main_base_dmg + weapon_baked.get("MAIN_DMG_RATING", 0)
        + wctx.main_dmg_rank_mod, is_h2h_pc=wctx.is_h2h)
    main_dmg_for_rank = (wctx.main_base_dmg + wctx.main_dmg_rank_mod)

    calc.section("Baked weapon / context constants")
    calc.put("wrank", "weapon rank", wrank)
    calc.put("smul", "STR->ATT mult", smul)
    calc.put("dmul", "DEX->ACC mult", dmul)
    calc.put("main_delay", "main delay", wctx.main_delay)
    calc.put("cap_main", "pDIF cap", cap_main)
    calc.put("main_hit_cap", "hit-rate cap", main_cap)
    calc.put("wsp", "WEAPON_SKILL_POWER", CONF.weapon_skill_power)
    calc.put("wb_attp", "weapon ATTP", weapon_baked.get("ATTP", 0))
    calc.put("wb_ftp", "weapon ANY_FTP_BONUS", weapon_baked.get("ANY_FTP_BONUS", 0))
    calc.put("wb_wsacc", "weapon WSACC", weapon_baked.get("WSACC", 0))

    calc.section("Gear deltas (candidate - equipped)")
    for m in TRACKED:
        calc.put(f"d_{m}", f"delta {m}",
                 f"=({gc_total[m]})-({eb_total[m]})")

    calc.section("Final candidate stats (anchored-delta)")
    for st in ("STR", "DEX", "VIT", "AGI", "INT", "MND", "CHR"):
        calc.put(f"{st}_c", f"{st}",
                 f"=({stat[st.lower() + '_a']})+({R['d_' + st]})")
    calc.put("attp_a", "ATTP (equipped)",
             f"=({R['wb_attp']})+({eb_total['ATTP']})")
    calc.put("attp_c", "ATTP (candidate)",
             f"=({R['wb_attp']})+({gc_total['ATTP']})")
    calc.put("core_a", "ATT core (anchor)",
             f"=ROUND({stat['att_a']}/(1+({R['attp_a']})/100),0)")
    calc.put("core_c", "ATT core (candidate)",
             f"=({R['core_a']})+({R['d_ATT']})"
             f"+INT(({R['STR_c']})*({R['smul']}))-INT(({stat['str_a']})*({R['smul']}))")
    calc.put("att_c", "ATT",
             f"=MAX(1,({R['core_c']})+INT(({R['core_c']})*({R['attp_c']})/100))")
    calc.put("acc_c", "ACC",
             f"=({stat['acc_a']})+({R['d_ACC']})"
             f"+INT(({R['DEX_c']})*({R['dmul']}))-INT(({stat['dex_a']})*({R['dmul']}))")
    calc.put("crit_rate_mod_c", "Crit rate mod",
             f"=({stat['crit_rate_a']})+({R['d_CRITHITRATE']})")
    calc.put("crit_dmg_c", "Crit dmg +%",
             f"=({stat['crit_dmg_a']})+({R['d_CRIT_DMG_INCREASE']})")
    calc.put("da_c", "Dbl Atk %", f"=({stat['da_a']})+({R['d_DOUBLE_ATTACK']})")
    calc.put("ta_c", "Trpl Atk %", f"=({stat['ta_a']})+({R['d_TRIPLE_ATTACK']})")
    calc.put("qa_c", "Quad Atk %", f"=({stat['qa_a']})+({R['d_QUAD_ATTACK']})")
    calc.put("hg_c", "Haste-Gear", f"=({stat['haste_gear_a']})+({R['d_HASTE_GEAR']})")
    calc.put("hm_c", "Haste-Magic", f"=({stat['haste_magic_a']})+({R['d_HASTE_MAGIC']})")
    calc.put("ha_c", "Haste-Ability", f"=({stat['haste_ability_a']})+({R['d_HASTE_ABILITY']})")
    calc.put("dw_c", "Dual Wield %", f"=({stat['dual_wield_a']})+({R['d_DUAL_WIELD']})")
    calc.put("stp_c", "Store TP", f"=({stat['store_tp_a']})+({R['d_STORETP']})")
    calc.put("tpb_c", "TP Bonus", f"=({stat['tp_bonus_a']})+({R['d_TP_BONUS']})")
    calc.put("wsd_all_c", "WS dmg all %",
             f"=({stat['ws_dmg_all_a']})+({R['d_ALL_WSDMG_ALL_HITS']})")
    calc.put("wsd_first_c", "WS dmg 1st %",
             f"=({stat['ws_dmg_first_a']})+({R['d_ALL_WSDMG_FIRST_HIT']})")
    calc.put("gear_ftp_c", "gear fTP (ANY_FTP/256)",
             f"=(({R['wb_ftp']})+({gc_total['ANY_FTP_BONUS']}))/256")
    calc.put("ws_acc_c", "WSACC",
             f"=({R['wb_wsacc']})+({gc_total['WSACC']})")
    calc.put("main_dmg_c", "weapon dmg",
             f"=({stat['wpn_dmg_a']})")
    calc.put("main_rank_dmg", "weapon dmg-for-rank", main_dmg_for_rank)

    calc.section("Target (selected)")
    calc.put("tgt_row", "target row",
             f"=MATCH({B_TARGET},list_targets,0)")
    for nm, col in (("tgt_def", 2), ("tgt_eva", 3), ("tgt_vit", 4),
                    ("tgt_agi", 5), ("tgt_crit_eva", 6), ("tgt_crit_def", 7)):
        cl = get_column_letter(col)
        calc.put(nm, nm, f"=INDEX(Targets!${cl}$2:${cl}${tgt_last},{R['tgt_row']})")

    calc.section("Weaponskill (selected)")
    calc.put("ws_row", "ws row", f"=MATCH({B_WS},list_ws,0)")

    def ws_col(name, col):
        cl = get_column_letter(col)
        calc.put(name, name, f"=INDEX(WSDB!${cl}$2:${cl}${ws_last},{R['ws_row']})")

    ws_col("ws_numhits", 2)
    ws_col("ws_ftp1", 3); ws_col("ws_ftp2", 4); ws_col("ws_ftp3", 5)
    for k, col in zip(("STR", "DEX", "VIT", "AGI", "INT", "MND", "CHR"), range(6, 13)):
        ws_col(f"ws_wsc_{k}", col)
    ws_col("ws_atk1", 13); ws_col("ws_atk2", 14); ws_col("ws_atk3", 15)
    ws_col("ws_ign1", 16); ws_col("ws_ign2", 17); ws_col("ws_ign3", 18)
    ws_col("ws_has_ign", 19)
    ws_col("ws_crit1", 20); ws_col("ws_crit2", 21); ws_col("ws_crit3", 22)
    ws_col("ws_has_crit", 23)
    ws_col("ws_multi_ftp", 24)
    ws_col("ws_id_sel", 25)
    # Per-WS WS-damage % (WEAPONSKILL_DAMAGE_BASE + ws_id), gear-only like fTP:
    # full candidate total (weapon baked + 14-slot), keyed to the selected WS.
    if wsdmg_ids:
        terms = "+".join(
            f"IF(({R['ws_id_sel']})={wid},"
            f"({weapon_baked[f'WSDMG_{wid}']})+({gc_total[f'WSDMG_{wid}']}),0)"
            for wid in wsdmg_ids)
        calc.put("ws_dmg_pct", "per-WS WS dmg %", "=" + terms)
    else:
        calc.put("ws_dmg_pct", "per-WS WS dmg %", 0)

    calc.section("Shared terms")
    # fSTR (main)
    calc.put("statdiff", "STR-VIT",
             f"=MAX(({(7 + wrank * 2) * -2}),MIN(({(14 + wrank * 2) * 2}),"
             f"({R['STR_c']})-({R['tgt_vit']})))")
    sd = R["statdiff"]
    calc.put("fbonus", "fSTR bracket",
             f"=IF({sd}>=12,4,IF({sd}>=6,6,IF({sd}>=1,7,IF({sd}>=-2,8,"
             f"IF({sd}>=-7,9,IF({sd}>=-15,10,IF({sd}>=-21,12,13)))))))")
    lower = -1 if wrank == 0 else -wrank
    calc.put("fstr_main", "fSTR",
             f"=MAX({lower},MIN({wrank + 8},(({sd})+({R['fbonus']}))/4))")
    # multi-attack
    calc.put("qf", "QA frac", f"=MIN(MAX(({R['qa_c']}),0),100)/100")
    calc.put("tf", "TA frac", f"=MIN(MAX(({R['ta_c']}),0),100)/100")
    calc.put("df", "DA frac", f"=MIN(MAX(({R['da_c']}),0),100)/100")
    calc.put("hits_swing", "hits/swing",
             f"=1+({R['qf']})*3+(1-({R['qf']}))*({R['tf']})*2"
             f"+(1-({R['qf']}))*(1-({R['tf']}))*({R['df']})")
    calc.put("per_roll", "extra hits/roll", f"=({R['hits_swing']})-1")
    calc.put("p_proc", "P(multi proc)",
             f"=1-(1-({R['qf']}))*(1-({R['tf']}))*(1-({R['df']}))")
    # crit
    calc.put("ddex", "DEX-AGI", f"=({R['DEX_c']})-({R['tgt_agi']})")
    dd = R["ddex"]
    calc.put("csb", "crit stat bonus",
             f"=IF({dd}>50,0.15,IF({dd}>=40,({dd}-35)/100,IF({dd}>=30,0.04,"
             f"IF({dd}>=20,0.03,IF({dd}>=14,0.02,IF({dd}>=7,0.01,0))))))")
    calc.put("crit_dmg_mult", "crit dmg mult",
             f"=(100+MAX(0,MIN(100,({R['crit_dmg_c']})-({R['tgt_crit_def']}))))/100")
    calc.put("crit_auto", "swing crit",
             f"=MAX(0.05,MIN(1,0.05+({R['csb']})+({R['crit_rate_mod_c']})/100"
             f"-({R['tgt_crit_eva']})/100))")

    calc.section("Auto-attack")
    pdif_auto = _emit_pdif(calc, "auto", R["att_c"], "1", "0", R["crit_auto"],
                           tgt_def=R["tgt_def"], crit_dmg_mult_ref=R["crit_dmg_mult"],
                           cap=cap_main)
    calc.put("hr_auto", "hit rate",
             "=" + _hit_rate(R["acc_c"], R["tgt_eva"], main_cap))
    calc.put("main_hit_dmg", "main hit dmg",
             f"=(({R['main_dmg_c']})+({R['fstr_main']}))*({pdif_auto})")
    calc.put("dmg_round", "dmg/round",
             f"=({R['hits_swing']})*({R['hr_auto']})*({R['main_hit_dmg']})")
    # swing interval (single-wield: no sub delay, dw_mult=1)
    wdelay = wctx.main_delay * 1000.0 / 60.0
    calc.put("hm_clamp", "haste-magic clamp",
             f"=MAX(-1,MIN(0.4375,({R['hm_c']})/10000))")
    calc.put("ha_clamp", "haste-ability clamp",
             f"=MAX(-0.25,MIN(0.25,({R['ha_c']})/10000))")
    calc.put("hg_clamp", "haste-gear clamp",
             f"=MAX(-0.25,MIN(0.25,({R['hg_c']})/10000))")
    calc.put("haste_mult", "haste mult",
             f"=MAX(0.2,MIN(2,1-({R['hm_clamp']})-({R['ha_clamp']})-({R['hg_clamp']})))")
    calc.put("interval_ms", "swing interval ms",
             f"=MAX({wdelay}*0.2,MIN({wdelay}*2,{wdelay}*({R['haste_mult']})))")
    calc.put("auto_dps", "AUTO DPS",
             f"=({R['dmg_round']})/(({R['interval_ms']})/1000)")
    # TP per round (single-wield: d = main delay, constant)
    tp_base = tp_return_pc(wctx.main_delay)
    calc.put("tp_per_hit", "TP/hit",
             f"={tp_base}*(1+({R['stp_c']})/100)")
    calc.put("tp_round", "TP/round",
             f"=({R['hits_swing']})*({R['hr_auto']})*({R['tp_per_hit']})")

    calc.section("Weaponskill")
    calc.put("tp_used", "TP used",
             f"=MAX(0,MIN(3000,{B_TP}+({R['tpb_c']})))")
    tu = R["tp_used"]
    calc.put("wsc", "WSC",
             "=" + "+".join(
                 f"INT(({R[s + '_c']})*({R['ws_wsc_' + s]}))"
                 for s in ("STR", "DEX", "VIT", "AGI", "INT", "MND", "CHR")))
    calc.put("main_base", "WS per-hit base",
             f"=INT(({R['main_dmg_c']})+({R['fstr_main']})+({R['wsc']}))")
    calc.put("ftp_first", "fTP first",
             "=" + _ftp(tu, R["ws_ftp1"], R["ws_ftp2"], R["ws_ftp3"])
             + f"+({R['gear_ftp_c']})")
    calc.put("ftp_rest", "fTP rest",
             f"=IF(({R['ws_multi_ftp']})=1,({R['ftp_first']}),1)")
    calc.put("ws_acc_bonus", "WS acc bonus",
             f"=CEILING(({R['gear_ftp_c']})*100,1)+({R['ws_acc_c']})")
    calc.put("ws_atk_mult", "WS atk mult",
             "=" + _ftp(tu, R["ws_atk1"], R["ws_atk2"], R["ws_atk3"]))
    calc.put("ws_ign", "WS def ignored",
             f"=IF(({R['ws_has_ign']})=1,"
             + _ftp(tu, R["ws_ign1"], R["ws_ign2"], R["ws_ign3"]) + ",0)")
    calc.put("ws_crit_factor", "WS crit fTP",
             "=" + _ftp(tu, R["ws_crit1"], R["ws_crit2"], R["ws_crit3"]))
    calc.put("ws_crit", "WS crit rate",
             f"=IF(({R['ws_has_crit']})=1,MAX(0.05,MIN(1,0.05+({R['csb']})"
             f"+({R['crit_rate_mod_c']})/100+({R['ws_crit_factor']})"
             f"-({R['tgt_crit_eva']})/100)),0)")
    pdif_ws = _emit_pdif(calc, "ws", R["att_c"], R["ws_atk_mult"], R["ws_ign"],
                         R["ws_crit"], tgt_def=R["tgt_def"],
                         crit_dmg_mult_ref=R["crit_dmg_mult"], cap=cap_main)
    calc.put("hr_first", "hit rate (1st)",
             "=" + _hit_rate(f"({R['acc_c']})+({R['ws_acc_bonus']})+100",
                             R["tgt_eva"], main_cap))
    calc.put("hr_ws", "hit rate (WS)",
             "=" + _hit_rate(f"({R['acc_c']})+({R['ws_acc_bonus']})",
                             R["tgt_eva"], main_cap))
    calc.put("first_dmg", "first hit dmg",
             f"=({R['hr_first']})*({R['main_base']})*({R['ftp_first']})*({pdif_ws})")
    calc.put("first_bonus", "first-hit bonus",
             f"=({R['first_dmg']})*({R['wsd_first_c']})/100")
    calc.put("rem_hits", "remaining main hits",
             f"=MAX(({R['ws_numhits']})-1,0)")
    calc.put("rem_main", "remaining main dmg",
             f"=({R['rem_hits']})*({R['hr_ws']})*({R['main_base']})"
             f"*({R['ftp_rest']})*({pdif_ws})")
    calc.put("multi_main", "WS multi extra hits",
             f"=IF(OR(({R['per_roll']})<=0,({R['p_proc']})<=0.000000001),0,"
             f"(({R['per_roll']})/({R['p_proc']}))"
             f"*MIN(({R['ws_numhits']})*({R['p_proc']}),2))")
    calc.put("main_multi", "WS multi dmg",
             f"=({R['multi_main']})*({R['hr_ws']})*({R['main_base']})"
             f"*({R['ftp_rest']})*({pdif_ws})")
    calc.put("ws_raw", "WS raw",
             f"=({R['first_dmg']})+({R['rem_main']})+({R['main_multi']})")
    calc.put("ws_final", "WS DAMAGE",
             f"=(({R['ws_raw']})*(100+({R['wsd_all_c']})+({R['ws_dmg_pct']}))/100"
             f"+({R['first_bonus']}))*({R['wsp']})")

    calc.section("Rotation")
    calc.put("rot_dps", "ROTATION DPS",
             f"=IF(OR(({R['tp_round']})<=0,({R['interval_ms']})<=0),({R['auto_dps']}),"
             f"({R['auto_dps']})+({R['ws_final']})"
             f"/(({B_TP}/({R['tp_round']}))*({R['interval_ms']})/1000))")


# ---------------------------------------------------------------------------
# Engine baseline (Python) for the all-"keep equipped" set, for side-by-side.
# ---------------------------------------------------------------------------
def _engine_baseline(base_cs, tgts, ws_list):
    default_target = tgts[0]
    default_ws = None
    best = -1.0
    for w in ws_list:
        try:
            d = engine.ws_damage(base_cs, default_target, w, 1000).damage
        except Exception:
            continue
        if d > best:
            best, default_ws = d, w
    auto = engine.auto_attack_dps(base_cs, default_target)
    wsr = engine.ws_damage(base_cs, default_target, default_ws, 1000) if default_ws else None
    rot = (engine.rotation_dps(base_cs, default_target, default_ws, 1000).total_dps
           if default_ws else auto.dps)
    return default_target, default_ws, {
        "auto": auto.dps,
        "ws": wsr.damage if wsr else 0.0,
        "rotation": rot,
    }


# ---------------------------------------------------------------------------
# Build sheet (the interactive surface)
# ---------------------------------------------------------------------------
def _write_build(ws, calc_ref, default_target, default_ws, eng_base,
                 anchor: Anchor) -> None:
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 30
    for c in "DEFGH":
        ws.column_dimensions[c].width = 13

    ws.cell(1, 1, "Gear / Augment DPS Optimizer").font = Font(bold=True, size=14)
    ws.cell(2, 1, "Pick an item + augment per slot. Numbers recompute live. "
                  "\"(keep equipped)\" = your current piece.").font = Font(italic=True)

    # selectors
    ws.cell(4, 1, "Target:").font = BOLD
    ws.cell(4, 2, default_target.label).fill = IN_FILL
    ws.cell(5, 1, "Weaponskill:").font = BOLD
    ws.cell(5, 2, default_ws.name if default_ws else "").fill = IN_FILL
    ws.cell(6, 1, "WS TP:").font = BOLD
    ws.cell(6, 2, 1000).fill = IN_FILL

    _add_dv(ws, "=list_targets", ["B4"])
    _add_dv(ws, "=list_ws", ["B5"])

    # output panel: live Calc-sheet refs alongside the Python engine baseline.
    ws.cell(8, 1, "Result").font = HDR_FONT
    ws.cell(8, 1).fill = HDR_FILL
    for col in (2, 3):
        ws.cell(8, col).fill = HDR_FILL
    ws.cell(8, 2, "live").font = HDR_FONT
    ws.cell(8, 3, "engine baseline").font = HDR_FONT
    out_rows = (
        (9, "Auto-attack DPS", calc_ref["auto_dps"], eng_base["auto"]),
        (10, "WS damage", calc_ref["ws_final"], eng_base["ws"]),
        (11, "Rotation DPS", calc_ref["rot_dps"], eng_base["rotation"]),
    )
    for r, label, live_ref, base in out_rows:
        ws.cell(r, 1, label).font = BOLD
        c = ws.cell(r, 2, f"={live_ref}")
        c.fill = OUT_FILL
        c.number_format = "0.0"
        ws.cell(r, 3, round(base, 1))
    ws.cell(12, 1, "(baseline = all slots \"keep equipped\", vs the default "
                   "target/WS above)").font = Font(italic=True, size=9)

    # per-slot table
    hdr = ["Slot", "Item", "Augment", "Count", "Free mod 1", "Val 1",
           "Free mod 2", "Val 2"]
    for j, h in enumerate(hdr, start=1):
        c = ws.cell(13, j, h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
    for i, s in enumerate(SWAP_SLOTS):
        r = SLOT_ROW0 + i
        ws.cell(r, 1, gear.SLOT_NAMES[s]).font = BOLD
        ws.cell(r, 2, "(keep equipped)").fill = IN_FILL
        ws.cell(r, 3, "(none)").fill = IN_FILL
        ws.cell(r, 4, 1).fill = IN_FILL
        ws.cell(r, 5, "(none)").fill = IN_FILL
        ws.cell(r, 7, "(none)").fill = IN_FILL
        _add_dv(ws, f"=list_slot{s}", [f"B{r}"])
        _add_dv(ws, "=list_augs", [f"C{r}"])
        _add_dv(ws, "=list_freemods", [f"E{r}", f"G{r}"])
        dv = DataValidation(type="whole", operator="between",
                            formula1="0", formula2="4", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"D{r}")

    ws.freeze_panes = "A14"


def _add_dv(ws, formula1: str, cells: list[str]) -> None:
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    ws.add_data_validation(dv)
    for c in cells:
        dv.add(c)


# ---------------------------------------------------------------------------
# Guide sheet
# ---------------------------------------------------------------------------
def _write_guide(ws, anchor: Anchor, ngear, naug, nws, ntgt) -> None:
    ws.column_dimensions["A"].width = 100
    lines = [
        ("Gear / Augment DPS Optimizer", True),
        ("", False),
        (f"Built for {anchor.main_job}{anchor.main_lvl}/{anchor.sub_job}"
         f"{anchor.sub_lvl}.  Curated gear: {ngear} items, {naug} augments, "
         f"{nws} weaponskills, {ntgt} targets.", False),
        ("", False),
        ("HOW TO USE", True),
        ("1. Go to the Build sheet.", False),
        ("2. Pick a Target and Weaponskill at the top, and set WS TP.", False),
        ("3. For each slot, choose an Item from the dropdown:", False),
        ("     (keep equipped) = your current piece (with its real augments)", False),
        ("     (none)          = nothing in that slot", False),
        ("     or any curated item the dropdown offers for that slot.", False),
        ("4. Optionally add an Augment (and Count 1-4 for stacked catalysts),", False),
        ("   or type a Free mod + value for an augment not in the catalog.", False),
        ("5. The Auto/WS/Rotation numbers on the Build sheet update instantly.", False),
        ("", False),
        ("ACCURACY", True),
        ("All damage formulas are the server's own (ported from the Lua/C++).", False),
        ("The model anchors to your /mystats numbers on the Stats sheet, then", False),
        ("applies gear deltas. Keep the Stats sheet matching YOUR unbuffed,", False),
        ("no-food /mystats for the weapon you actually wield.", False),
        ("The 'engine baseline' column on Build shows the exact engine numbers", False),
        ("for the all-keep-equipped set so you can confirm the sheet agrees.", False),
        ("", False),
        ("NOTES", True),
        ("- The weapon is fixed. To compare weapons, re-generate with that", False),
        ("  weapon equipped and a fresh /mystats anchor.", False),
        ("- 'keep equipped' uses your real augments; picking the same item by", False),
        ("  name instead gives its un-augmented base so you can re-augment it.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(i, 1, text)
        if bold:
            c.font = Font(bold=True, size=12)
